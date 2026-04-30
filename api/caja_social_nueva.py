import json, base64, math, unicodedata
from datetime import datetime, date as date_type
from io import BytesIO
from http.server import BaseHTTPRequestHandler
from collections import Counter

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.filters import FilterColumn, Filters

# ── Constants ─────────────────────────────────────────────────────────────────
GRAVAMEN   = "GRAVAMEN MOVS FINANCIEROS"
DCTOS      = "DCTOS DE NOMINA"
ESPECIALES = {GRAVAMEN, DCTOS}

C_GREEN  = "FFE8F5E9"
C_YELLOW = "FFFFFCCC"
C_RED    = "FFFFEBEE"

_side  = Side(style="thin", color="FF000000")
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False):
    return Font(name="Trebuchet MS", bold=bold, size=12)


# ── Date helpers ──────────────────────────────────────────────────────────────
def _to_date(val):
    """Convert 'DD/MM/YYYY' string or openpyxl date/datetime to date object."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date_type):
        return val
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_rango(fecha_ini_str, fecha_fin_str):
    ini = datetime.strptime(fecha_ini_str, "%Y-%m-%d").date()
    fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
    return ini, fin


def _en_rango(fecha_val, ini, fin):
    d = _to_date(fecha_val)
    return d is not None and ini <= d <= fin


# ── Amount parsing ────────────────────────────────────────────────────────────
def parse_monto(s):
    """Parse Colombian-format string '1.234.567,00' or numeric value to int."""
    if s is None or s == "":
        return 0
    if isinstance(s, (int, float)):
        try:
            v = float(s)
            return 0 if math.isnan(v) else int(v)
        except Exception:
            return 0
    cleaned = str(s).strip().replace(".", "").replace(",", ".")
    try:
        return int(float(cleaned))
    except Exception:
        return 0


# ── Bank parsing ──────────────────────────────────────────────────────────────
def leer_banco(data_b64, ini, fin):
    """Return (debito_rows, credito_rows) filtered to [ini, fin].
    Each row tuple: (fecha_str, desc, doc, amount, info)
    """
    wb = openpyxl.load_workbook(BytesIO(base64.b64decode(data_b64)), data_only=True)
    ws = wb.active

    header_found = False
    fecha_i = desc_i = doc_i = deb_i = cred_i = info_i = None

    debito_rows  = []
    credito_rows = []

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            vals = [str(c).lower() if c else "" for c in row]
            # Detect actual data header (contains both 'fecha' and 'bito' from Débito)
            if any("fecha" in v for v in vals) and any("bito" in v for v in vals):
                header_found = True
                for i, v in enumerate(vals):
                    if "fecha" in v and fecha_i is None:
                        fecha_i = i
                    elif "escripci" in v and desc_i is None:
                        desc_i = i
                    elif "ocumento" in v and doc_i is None:
                        doc_i = i
                    elif "bito" in v and "cr" not in v and deb_i is None:
                        deb_i = i
                    elif "r" in v and "dito" in v and cred_i is None:
                        cred_i = i
                    elif "nformaci" in v and info_i is None:
                        info_i = i
            continue

        if fecha_i is None:
            continue

        fecha    = row[fecha_i]  if fecha_i  is not None else None
        desc     = row[desc_i]   if desc_i   is not None else None
        doc      = row[doc_i]    if doc_i    is not None else None
        deb_raw  = row[deb_i]    if deb_i    is not None else None
        cred_raw = row[cred_i]   if cred_i   is not None else None
        info     = row[info_i]   if info_i   is not None else None

        if fecha is None and desc is None:
            continue

        # Date filter
        if not _en_rango(fecha, ini, fin):
            continue

        deb_amt  = parse_monto(deb_raw)
        cred_amt = parse_monto(cred_raw)

        fecha_s = str(fecha) if fecha else ""
        desc_s  = str(desc)  if desc  else ""
        doc_s   = str(doc)   if doc   else ""
        info_s  = str(info)  if info  else ""

        if deb_amt > 0:
            debito_rows.append((fecha_s, desc_s, doc_s, deb_amt, info_s))
        elif cred_amt > 0:
            credito_rows.append((fecha_s, desc_s, doc_s, cred_amt, info_s))

    return debito_rows, credito_rows


# ── Siigo parsing ─────────────────────────────────────────────────────────────
def leer_siigo(data_b64, ini, fin):
    """Return (siigo_credito, siigo_debito) filtered to [ini, fin].
    Each entry tuple: (comp, fecha_str, ident, amount)
    siigo_credito → cross-matches CS Débito
    siigo_debito  → cross-matches CS Crédito
    """
    wb = openpyxl.load_workbook(BytesIO(base64.b64decode(data_b64)), data_only=True)
    ws = wb.active

    header_found  = False
    siigo_credito = []
    siigo_debito  = []

    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[2] is not None and str(row[2]).strip() == "Comprobante":
                header_found = True
            continue

        comp = row[2]
        if not comp:
            continue
        comp_s = str(comp).strip()
        if not comp_s:
            continue

        fecha    = row[4]
        ident    = row[5]
        deb_amt  = parse_monto(row[12])
        cred_amt = parse_monto(row[13])

        # Date filter
        if not _en_rango(fecha, ini, fin):
            continue

        fecha_s = str(fecha) if fecha else ""
        ident_s = str(ident) if ident else ""

        if deb_amt > 0:
            siigo_debito.append((comp_s, fecha_s, ident_s, deb_amt))
        if cred_amt > 0:
            siigo_credito.append((comp_s, fecha_s, ident_s, cred_amt))

    return siigo_credito, siigo_debito


# ── Matching ──────────────────────────────────────────────────────────────────
def match_multiset(bank_rows, siigo_entries):
    """Multiset match by amount.
    Returns bank_matched (set of indices) and siigo_only (unmatched siigo entries).
    """
    siigo_counter = Counter(e[3] for e in siigo_entries)

    bank_matched = set()
    for i, row in enumerate(bank_rows):
        amt = row[3]
        if siigo_counter.get(amt, 0) > 0:
            siigo_counter[amt] -= 1
            bank_matched.add(i)

    remaining = dict(siigo_counter)
    siigo_only = []
    for entry in siigo_entries:
        amt = entry[3]
        if remaining.get(amt, 0) > 0:
            remaining[amt] -= 1
            siigo_only.append(entry)

    return bank_matched, siigo_only


# ── Excel generation ──────────────────────────────────────────────────────────
def _write_data_row(ws, row_idx, values, fill_color, hidden=False):
    fill = _fill(fill_color)
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font   = _font()
        cell.border = BORDER
        cell.fill   = fill
        if col_idx == 4 and isinstance(val, (int, float)):
            cell.number_format = "#,##0"
    if hidden:
        ws.row_dimensions[row_idx].hidden = True


def _write_header(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = Font(name="Trebuchet MS", bold=True, color="FFFFFFFF", size=12)
        cell.fill   = PatternFill("solid", fgColor="FF4472C4")
        cell.border = BORDER


def _set_col_widths(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def _apply_filter(ws, last_data_row, col_count, visible_descs=None):
    """Set AutoFilter ref and optionally a Description column filter."""
    col_letter = chr(ord("A") + col_count - 1)
    ws.auto_filter.ref = f"A1:{col_letter}{last_data_row}"
    if visible_descs:
        vals = [v for v in sorted(visible_descs) if v]
        fc = FilterColumn(colId=1)   # column B = index 1 (0-based from A)
        fc.filters = Filters(filter=vals)
        ws.auto_filter.filterColumn.append(fc)


def generar_excel(debito_rows, credito_rows, siigo_credito, siigo_debito):
    deb_matched,  siigo_cred_only = match_multiset(debito_rows,  siigo_credito)
    cred_matched, siigo_deb_only  = match_multiset(credito_rows, siigo_debito)

    wb = openpyxl.Workbook()

    # ── DEBITO sheet ──────────────────────────────────────────────────────────
    ws_deb = wb.active
    ws_deb.title = "DEBITO"
    _set_col_widths(ws_deb, {"A": 14, "B": 35, "C": 15, "D": 18, "E": 45})
    _write_header(ws_deb, ["Fecha", "Descripción", "Documento", "-Débito", "Información Adicional"])

    # Classify rows
    green_deb, yellow_deb, red_deb = [], [], []
    especiales_rows = []   # for "Para Nataly" sheet
    for orig_i, (fecha, desc, doc, amt, info) in enumerate(debito_rows):
        is_especial = desc in ESPECIALES
        is_matched  = orig_i in deb_matched
        entry = (fecha, desc, doc, amt, info, is_especial)
        if is_especial:
            red_deb.append(entry)
            especiales_rows.append((fecha, desc, doc, amt, info))
        elif is_matched:
            green_deb.append(entry)
        else:
            red_deb.append(entry)

    for comp, fecha, ident, amt in siigo_cred_only:
        yellow_deb.append((fecha, comp, ident, amt, "", False))

    # Sort within each group by amount desc
    key_amt = lambda r: -r[3]
    green_deb.sort(key=key_amt)
    yellow_deb.sort(key=key_amt)
    red_deb.sort(key=key_amt)

    total_gravamen = 0
    total_dctos    = 0
    visible_descs_deb = set()
    row_idx = 2

    for rows, color in [(green_deb, C_GREEN), (yellow_deb, C_YELLOW), (red_deb, C_RED)]:
        for fecha, desc, doc, amt, info, is_especial in rows:
            hidden = is_especial
            _write_data_row(ws_deb, row_idx, [fecha, desc, doc, amt, info], color, hidden)
            if is_especial:
                if desc == GRAVAMEN:
                    total_gravamen += amt
                else:
                    total_dctos += amt
            else:
                visible_descs_deb.add(str(desc) if desc else "")
            row_idx += 1

    last_data_deb = row_idx - 1
    _apply_filter(ws_deb, last_data_deb, 5, visible_descs_deb)

    # Blank separator + total rows
    row_idx += 1
    for label, total in [(GRAVAMEN, total_gravamen), (DCTOS, total_dctos)]:
        for col_idx, val in [(2, label), (4, total)]:
            cell = ws_deb.cell(row=row_idx, column=col_idx, value=val)
            cell.font   = _font(bold=True)
            cell.border = BORDER
            if col_idx == 4:
                cell.number_format = "#,##0"
        row_idx += 1

    # ── CREDITO sheet ─────────────────────────────────────────────────────────
    ws_cred = wb.create_sheet("CREDITO")
    _set_col_widths(ws_cred, {"A": 14, "B": 35, "C": 15, "D": 18, "E": 45})
    _write_header(ws_cred, ["Fecha", "Descripción", "Documento", "+Crédito", "Información Adicional"])

    green_cred, yellow_cred, red_cred = [], [], []
    for orig_i, row in enumerate(credito_rows):
        if orig_i in cred_matched:
            green_cred.append(row)
        else:
            red_cred.append(row)
    for comp, fecha, ident, amt in siigo_deb_only:
        yellow_cred.append((fecha, comp, ident, amt, ""))

    green_cred.sort(key=key_amt)
    yellow_cred.sort(key=key_amt)
    red_cred.sort(key=key_amt)

    row_idx_c = 2
    for rows, color in [(green_cred, C_GREEN), (yellow_cred, C_YELLOW), (red_cred, C_RED)]:
        for row in rows:
            _write_data_row(ws_cred, row_idx_c, list(row), color)
            row_idx_c += 1

    last_data_cred = row_idx_c - 1
    _apply_filter(ws_cred, last_data_cred, 5)

    # ── Para Nataly sheet ─────────────────────────────────────────────────────
    ws_nat = wb.create_sheet("Para Nataly")
    _set_col_widths(ws_nat, {"A": 14, "B": 35, "C": 15, "D": 18, "E": 45})
    _write_header(ws_nat, ["Fecha", "Descripción", "Documento", "-Débito", "Información Adicional"])

    # GRAVAMEN rows then DCTOS rows, each sorted by amount desc
    gravamen_rows = sorted(
        [r for r in especiales_rows if r[1] == GRAVAMEN], key=key_amt)
    dctos_rows = sorted(
        [r for r in especiales_rows if r[1] == DCTOS], key=key_amt)

    row_idx_n = 2
    for row in gravamen_rows + dctos_rows:
        fecha, desc, doc, amt, info = row
        _write_data_row(ws_nat, row_idx_n, [fecha, desc, doc, amt, info], C_RED)
        row_idx_n += 1

    last_data_nat = row_idx_n - 1
    _apply_filter(ws_nat, last_data_nat, 5)

    # Total rows
    row_idx_n += 1
    for label, total in [(GRAVAMEN, total_gravamen), (DCTOS, total_dctos)]:
        for col_idx, val in [(2, label), (4, total)]:
            cell = ws_nat.cell(row=row_idx_n, column=col_idx, value=val)
            cell.font   = _font(bold=True)
            cell.border = BORDER
            if col_idx == 4:
                cell.number_format = "#,##0"
        row_idx_n += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Totals ────────────────────────────────────────────────────────────────────
def calcular_resumen(debito_rows, credito_rows, siigo_credito, siigo_debito):
    total_cs_deb   = sum(r[3] for r in debito_rows)
    total_sii_cred = sum(e[3] for e in siigo_credito)
    total_cs_cred  = sum(r[3] for r in credito_rows)
    total_sii_deb  = sum(e[3] for e in siigo_debito)
    diff_deb  = total_cs_deb  - total_sii_cred
    diff_cred = total_cs_cred - total_sii_deb

    return {
        "total_deb_banco":  total_cs_deb,
        "total_deb_siigo":  total_sii_cred,
        "diff_deb":         diff_deb,
        "total_cred_banco": total_cs_cred,
        "total_cred_siigo": total_sii_deb,
        "diff_cred":        diff_cred,
        "conciliado":       diff_deb == 0 and diff_cred == 0,
    }


# ── Validación (antes de procesar) ───────────────────────────────────────────
def _norm_csn(s):
    if s is None:
        return ''
    n = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return n.lower().strip()


def _validar_banco_csn(data_b64):
    """Valida que el archivo del banco tenga la fila de encabezados con las
    columnas críticas. Lanza ValueError con mensaje descriptivo si algo falta."""
    wb = openpyxl.load_workbook(BytesIO(base64.b64decode(data_b64)), data_only=True)
    ws = wb.active

    header_found = False
    fecha_i = deb_i = cred_i = desc_i = None

    for row in ws.iter_rows(values_only=True):
        vals = [str(c).lower() if c else '' for c in row]
        if any('fecha' in v for v in vals) and any('bito' in v for v in vals):
            header_found = True
            for i, v in enumerate(vals):
                if 'fecha' in v and fecha_i is None:
                    fecha_i = i
                elif 'escripci' in v and desc_i is None:
                    desc_i = i
                elif 'bito' in v and 'cr' not in v and deb_i is None:
                    deb_i = i
                elif 'r' in v and 'dito' in v and cred_i is None:
                    cred_i = i
            encontradas = [str(c) for c in row if c is not None]
            break

    if not header_found:
        raise ValueError(
            'Banco Caja Social: No se encontró la fila de encabezados. '
            'Verificá que el archivo es el extracto correcto descargado de Caja Social.'
        )

    faltantes = []
    if fecha_i is None:
        faltantes.append('Fecha')
    if deb_i is None:
        faltantes.append('Débito')
    if cred_i is None:
        faltantes.append('Crédito')

    if faltantes:
        raise ValueError(
            f'Banco Caja Social: Columnas no encontradas: {faltantes}. '
            f'Encabezados disponibles: {encontradas}'
        )


def _validar_siigo_csn(data_b64):
    """Valida que el archivo de Siigo tenga 'Comprobante' en col C y columnas
    de Débito/Crédito en las posiciones esperadas (cols 13 y 14)."""
    wb = openpyxl.load_workbook(BytesIO(base64.b64decode(data_b64)), data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 3).value
        if val and str(val).strip() == 'Comprobante':
            header_row = r
            break
    if header_row is None:
        raise ValueError(
            "Siigo: No se encontró 'Comprobante' en columna C. "
            "Verificá que el archivo es el reporte correcto exportado desde Siigo."
        )

    num_cols = ws.max_column
    if num_cols < 14:
        raise ValueError(
            f'Siigo: Se esperan al menos 14 columnas, el archivo tiene {num_cols}. '
            f'Verificá que es el reporte de Siigo completo.'
        )

    h13 = ws.cell(header_row, 13).value  # col M — debe ser Débito
    h14 = ws.cell(header_row, 14).value  # col N — debe ser Crédito
    errores = []
    if 'deb' not in _norm_csn(h13):
        errores.append(f'col M (13) tiene "{h13}", se esperaba "Débito"')
    if 'cred' not in _norm_csn(h14):
        errores.append(f'col N (14) tiene "{h14}", se esperaba "Crédito"')

    if errores:
        disponibles = [ws.cell(header_row, c).value for c in range(1, num_cols + 1)]
        raise ValueError(
            f'Siigo: Columnas de montos inesperadas — {"; ".join(errores)}. '
            f'Encabezados disponibles: {disponibles}'
        )


# ── HTTP Handler ──────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body   = json.loads(self.rfile.read(length))

        try:
            ini, fin = _parse_rango(body["fecha_inicio"], body["fecha_fin"])

            _validar_banco_csn(body["banco_b64"])
            _validar_siigo_csn(body["siigo_b64"])

            debito_rows, credito_rows   = leer_banco(body["banco_b64"], ini, fin)
            siigo_credito, siigo_debito = leer_siigo(body["siigo_b64"], ini, fin)

            excel_bytes = generar_excel(debito_rows, credito_rows, siigo_credito, siigo_debito)
            resumen     = calcular_resumen(debito_rows, credito_rows, siigo_credito, siigo_debito)

            payload = {
                "archivo_b64": base64.b64encode(excel_bytes).decode(),
                "nombre":      "conciliacion_caja_social_nueva.xlsx",
                "resumen":     resumen,
            }
            self._respond(200, payload)

        except Exception as exc:
            import traceback
            self._respond(500, {"error": str(exc), "detail": traceback.format_exc()})

    def _respond(self, status, data):
        resp = json.dumps(data).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(resp)))
        self.end_headers()
        self.wfile.write(resp)

    def log_message(self, *args):
        pass
