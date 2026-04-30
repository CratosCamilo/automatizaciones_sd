"""
Vercel Python serverless function — Conciliación DIAN vs Siigo.
Recibe los archivos en base64 vía JSON POST y devuelve el Excel procesado.
"""

from http.server import BaseHTTPRequestHandler
import base64
import io
import json
import zipfile

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import pandas as pd


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

DIAN_KEEP_LEFT = [
    'Tipo de documento',
    'CUFE/CUDE',
    'Folio',
    'Prefijo',
    'Fecha Emisión',
    'NIT Emisor',
    'Nombre Emisor',
]

SIIGO_REMOVE    = {'Sucursal', 'Base gravada', 'Base exenta'}
COLS_MONETARIAS = {'IVA', 'ICUI', 'Total', 'Diferencia IVA'}
FMT_MILES       = '#,##0.##'
FUENTE          = 'Trebuchet MS'
OUTPUT_NAME     = 'CONCILIACION_FACTURAS.xlsx'

COLOR_HEADER_BG = '4472C4'
COLOR_HEADER_FG = 'FFFFFF'
COLOR_FALTA_BG  = 'F2DCDB'
COLOR_EXTRA_BG  = 'FDE9D9'
COLOR_OK_BG     = 'EBF1DE'
COLOR_TEXTO     = '000000'
COLOR_BORDE     = 'BFBFBF'


# ─────────────────────────────────────────────────────────────────────────────
# FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────

def _clean(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _dian_key(row):
    pref = str(row['Prefijo']).strip() if row['Prefijo'] not in (None, '') else ''
    try:
        folio = str(int(float(row['Folio'])))
    except (ValueError, TypeError):
        folio = str(row['Folio']).strip() if row['Folio'] not in (None, '') else ''
    return f'{pref}-{folio}'


def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _borde():
    s = Side(style='thin', color=COLOR_BORDE)
    return Border(left=s, right=s, top=s, bottom=s)


# ─────────────────────────────────────────────────────────────────────────────
# VALIDACIÓN (se ejecuta antes de procesar; lanza ValueError si algo falta)
# ─────────────────────────────────────────────────────────────────────────────

DIAN_CRITICAS  = ['Prefijo', 'Folio', 'IVA', 'Total']
SIIGO_CRITICAS = ['Comprobante', 'Factura proveedor', 'Total', 'Nombre tercero', 'IVA']


def _validar_dian(zip_bytes):
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith('.xlsx')]
        if not xlsx_names:
            raise ValueError('No se encontró ningún .xlsx dentro del ZIP de la DIAN')
        xlsx_data = zf.read(xlsx_names[0])

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else ''
               for c in next(ws.iter_rows(min_row=1, max_row=1))]
    wb.close()

    faltantes = [c for c in DIAN_CRITICAS if c not in headers]
    if faltantes:
        disponibles = [h for h in headers if h]
        raise ValueError(
            f'DIAN: Columnas críticas no encontradas: {faltantes}. '
            f'Disponibles: {disponibles}'
        )


def _validar_siigo(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    all_rows = list(ws.values)

    header_idx = None
    for i, row in enumerate(all_rows):
        if row and str(row[0]).strip() == 'Comprobante':
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(
            "Siigo: No se encontró 'Comprobante' como encabezado. "
            "Verificá que el archivo es el reporte correcto exportado desde Siigo."
        )

    headers = [str(h).strip() if h is not None else '' for h in all_rows[header_idx]]
    faltantes = [c for c in SIIGO_CRITICAS if c not in headers]
    if faltantes:
        disponibles = [h for h in headers if h]
        raise ValueError(
            f'Siigo: Columnas críticas no encontradas: {faltantes}. '
            f'Disponibles: {disponibles}'
        )


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO DIAN
# ─────────────────────────────────────────────────────────────────────────────

def procesar_dian(zip_bytes):
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith('.xlsx')]
        if not xlsx_names:
            raise ValueError('No se encontró ningún .xlsx dentro del ZIP de la DIAN')
        xlsx_data = zf.read(xlsx_names[0])

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
    ws = wb.active
    sheet_name = ws.title

    headers = [c.value for c in ws[1]]
    rows    = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    df = pd.DataFrame(rows, columns=headers)

    if 'IVA' not in headers or 'Total' not in headers:
        raise ValueError(f'El archivo DIAN no tiene las columnas esperadas. Columnas: {headers}')

    iva_idx   = headers.index('IVA')
    total_idx = headers.index('Total')

    keep_left = [c for c in DIAN_KEEP_LEFT if c in headers]

    keep_mid = []
    for col in headers[iva_idx : total_idx + 1]:
        if col in ('IVA', 'Total'):
            keep_mid.append(col)
        else:
            nums = pd.to_numeric(df[col], errors='coerce').fillna(0)
            if (nums != 0).any():
                keep_mid.append(col)

    df = df[keep_left + keep_mid]
    df['Total'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df = df[df['Total'] != 0].reset_index(drop=True)

    return df, sheet_name


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO SIIGO
# ─────────────────────────────────────────────────────────────────────────────

def procesar_siigo(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    all_rows = list(ws.values)

    header_idx = None
    for i, row in enumerate(all_rows):
        if row and str(row[0]).strip() == 'Comprobante':
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("No se encontró 'Comprobante' como encabezado en el archivo Siigo")

    headers = list(all_rows[header_idx])
    rows    = [list(r) for r in all_rows[header_idx + 1:]]
    df = pd.DataFrame(rows, columns=headers)

    df = df[[c for c in df.columns if c not in SIIGO_REMOVE]]
    df = df[df['Nombre tercero'].notna()]
    df = df[~df['Comprobante'].fillna('').str.upper().str.startswith('DS')]

    df['Total'] = pd.to_numeric(df['Total'], errors='coerce').fillna(0)
    df = df[df['Total'] != 0].reset_index(drop=True)

    return df


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DEL EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def generar_output(df_dian, df_siigo, sheet_name):
    siigo_cols    = list(df_siigo.columns)
    siigo_ordered = ['Total'] + [c for c in siigo_cols if c != 'Total']

    siigo_lookup = {}
    for _, row in df_siigo.iterrows():
        fp = str(row.get('Factura proveedor', '') or '').strip()
        if fp:
            siigo_lookup[fp] = row

    dian_cols = list(df_dian.columns)
    df_w = df_dian.copy()
    df_w['_key']     = df_w.apply(_dian_key, axis=1)
    df_w['_matched'] = df_w['_key'].isin(siigo_lookup)

    dian_only = df_w[~df_w['_matched']].sort_values('Total', ascending=False)
    matched   = df_w[ df_w['_matched']].sort_values('Total', ascending=False)

    dian_keys   = set(df_w['_key'])
    siigo_extra = df_siigo[
        ~df_siigo['Factura proveedor'].fillna('').str.strip().isin(dian_keys)
    ].sort_values('Total', ascending=False)

    n_dian_only  = len(dian_only)
    n_siigo_only = len(siigo_extra)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    n_dian      = len(dian_cols)
    diff_col    = n_dian + 1
    siigo_start = diff_col + 1

    dian_iva_col  = dian_cols.index('IVA') + 1 if 'IVA' in dian_cols else None
    siigo_iva_pos = siigo_ordered.index('IVA') if 'IVA' in siigo_ordered else None
    dian_iva_letter  = get_column_letter(dian_iva_col) if dian_iva_col else None
    siigo_iva_letter = (get_column_letter(siigo_start + siigo_iva_pos)
                        if siigo_iva_pos is not None else None)

    for i, col in enumerate(dian_cols, start=1):
        ws.cell(row=1, column=i, value=col)
    ws.cell(row=1, column=diff_col, value='Diferencia IVA')
    for i, col in enumerate(siigo_ordered, start=siigo_start):
        ws.cell(row=1, column=i, value=col)

    total_cols = ws.max_column
    r = 2

    for _, row in dian_only.iterrows():
        for c, col in enumerate(dian_cols, start=1):
            ws.cell(row=r, column=c, value=_clean(row[col]))
        r += 1

    for _, row in siigo_extra.iterrows():
        for c_off, col in enumerate(siigo_ordered):
            ws.cell(row=r, column=siigo_start + c_off, value=_clean(row[col]))
        r += 1

    for _, row in matched.iterrows():
        for c, col in enumerate(dian_cols, start=1):
            ws.cell(row=r, column=c, value=_clean(row[col]))
        if dian_iva_letter and siigo_iva_letter:
            ws.cell(row=r, column=diff_col,
                    value=f'={dian_iva_letter}{r}-{siigo_iva_letter}{r}')
        key = row['_key']
        if key in siigo_lookup:
            sr = siigo_lookup[key]
            for c_off, col in enumerate(siigo_ordered):
                ws.cell(row=r, column=siigo_start + c_off, value=_clean(sr[col]))
        r += 1

    # ── Estilos ───────────────────────────────────────────────────────────
    borde = _borde()
    mon_cols = {
        c for c in range(1, total_cols + 1)
        if ws.cell(row=1, column=c).value in COLS_MONETARIAS
    }
    mon_cols.add(diff_col)

    for c in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill      = _fill(COLOR_HEADER_BG)
        cell.font      = Font(name=FUENTE, bold=True, color=COLOR_HEADER_FG)
        cell.border    = borde
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    fila_inicio_extra = 2 + n_dian_only
    fila_inicio_ok    = fila_inicio_extra + n_siigo_only

    for data_r in range(2, ws.max_row + 1):
        if data_r < fila_inicio_extra:
            bg = COLOR_FALTA_BG
        elif data_r < fila_inicio_ok:
            bg = COLOR_EXTRA_BG
        else:
            bg = COLOR_OK_BG

        fill = _fill(bg)
        fnt  = Font(name=FUENTE, color=COLOR_TEXTO)

        for c in range(1, total_cols + 1):
            cell = ws.cell(row=data_r, column=c)
            cell.fill   = fill
            cell.font   = fnt
            cell.border = borde
            if c in mon_cols:
                cell.number_format = FMT_MILES

    all_headers = [ws.cell(row=1, column=c).value for c in range(1, total_cols + 1)]
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8
        )
        ws.column_dimensions[col_letter].width = max(min(max_len + 2, 45), 10)

    if 'CUFE/CUDE' in all_headers:
        cufe_col = get_column_letter(all_headers.index('CUFE/CUDE') + 1)
        ws.column_dimensions[cufe_col].width = 18

    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), n_dian_only, n_siigo_only


# ─────────────────────────────────────────────────────────────────────────────
# HANDLER HTTP (Vercel serverless)
# ─────────────────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass  # Silenciar logs por defecto

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
            data   = json.loads(body)

            if 'dian' not in data or 'siigo' not in data:
                return self._error(400, 'Se requieren los campos "dian" y "siigo" en base64')

            zip_bytes  = base64.b64decode(data['dian'])
            xlsx_bytes = base64.b64decode(data['siigo'])

            _validar_dian(zip_bytes)
            _validar_siigo(xlsx_bytes)

            df_dian, sheet_name = procesar_dian(zip_bytes)
            df_siigo = procesar_siigo(xlsx_bytes)
            output_bytes, n_falta, n_extra = generar_output(df_dian, df_siigo, sheet_name)

            n_ok = len(df_dian) - n_falta

            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition',
                f'attachment; filename="{OUTPUT_NAME}"')
            self.send_header('X-Falta', str(n_falta))
            self.send_header('X-Extra', str(n_extra))
            self.send_header('X-Ok',    str(n_ok))
            self._cors_headers()
            self.send_header('Access-Control-Expose-Headers', 'X-Falta, X-Extra, X-Ok')
            self.end_headers()
            self.wfile.write(output_bytes)

        except json.JSONDecodeError:
            self._error(400, 'El cuerpo de la solicitud no es JSON válido')
        except Exception as e:
            self._error(500, str(e))

    def _error(self, code, message):
        body = json.dumps({'error': message}).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self._cors_headers()
        self.end_headers()
        self.wfile.write(body)
