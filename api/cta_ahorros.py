from http.server import BaseHTTPRequestHandler
import json, base64, io, unicodedata
from datetime import datetime, date
from collections import defaultdict

# Descripciones de impuestos — mismo orden en que se muestran al final de CREDITO.
# Caja Social ya no entrega "Tipo Transacción" (ex N005/N328/N023/N467/N001):
# ahora la única forma de identificar estos impuestos es por la descripción.
IMPTOS_DESCS = [
    "DESC COMISION POR VENTAS T-DEB",
    "GRAVAMEN MOVS FINANCIEROS",
    "ND COMISION ADQUIRENCIA TC",
    "RETEFUENTE VTAS TARCREDIT",
    "NOTA DEBITO RETEICA",
]
IMPTOS_SET = set(IMPTOS_DESCS)

# Encabezados que mostramos en Hoja1 (se mantienen iguales al formato anterior
# para que el output siga siendo familiar). Las columnas que el nuevo extracto
# ya no trae (Saldo, Regional/Oficina, Tipo Transacción, Ref. Titular Cuenta)
# quedan vacías.
HOJA1_HEADERS = [
    "Fecha Transacción", "Descripción", "Valor", "Saldo",
    "Regional / Oficina", "Tipo Transacción", "Oficina",
    "Ref. Titular Cuenta", "Detalles Adicionales",
]

BANCO_SHEET = "AccountMovementsExtended"


# ── DETECCIÓN DE FORMATO ──────────────────────────────────────────────────────
def _es_banco_xlsx(data: bytes) -> bool:
    """Detecta el extracto de Caja Social por el nombre de la hoja interna."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        return BANCO_SHEET in wb.sheetnames
    except Exception:
        return False


def detectar_archivos(bytes_a: bytes, bytes_b: bytes):
    a_banco = _es_banco_xlsx(bytes_a)
    b_banco = _es_banco_xlsx(bytes_b)
    if a_banco and not b_banco:
        return bytes_a, bytes_b
    if b_banco and not a_banco:
        return bytes_b, bytes_a
    raise ValueError(
        "No se pudo identificar cuál archivo es el banco y cuál es Siigo. "
        "El extracto de Caja Social debe ser un .xlsx con la hoja 'AccountMovementsExtended'."
    )


# ── HELPERS ───────────────────────────────────────────────────────────────────
def parse_fecha(s):
    """Acepta string 'dd/mm/yyyy' o datetime/date."""
    if s is None or s == "":
        return None
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_monto_co(v):
    """Parsea monto en formato colombiano '1.234.567,00' (string) o numérico → float.
    Devuelve 0.0 si está vacío o no se puede parsear."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    # Colombiano: puntos = miles, coma = decimales
    cleaned = s.replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def _to_num_maybe(v):
    """Devuelve int si es entero, float si tiene decimales, '' si está vacío."""
    f = parse_monto_co(v)
    if v is None or v == "":
        return ""
    if f == 0 and (isinstance(v, str) and v.strip() == ""):
        return ""
    return int(f) if f == int(f) else f


def _localizar_header_banco(ws):
    """Encuentra la fila de encabezado del extracto y mapea las columnas.
    Devuelve (header_row, {fecha,desc,doc,deb,cred,oficina,info}) — índices 1-based.
    """
    for r in range(1, min(ws.max_row, 50) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        normed = [_norm_s(v) for v in vals]
        if any("fecha" in v for v in normed) and any("bito" in v and "cr" not in v for v in normed):
            cols = {}
            for i, v in enumerate(normed, 1):
                if "fecha" in v and "fecha" not in cols:
                    cols["fecha"] = i
                elif "escripci" in v and "desc" not in cols:
                    cols["desc"] = i
                elif "ocumento" in v and "doc" not in cols:
                    cols["doc"] = i
                elif "bito" in v and "cr" not in v and "deb" not in cols:
                    cols["deb"] = i
                elif "dito" in v and "cr" in v and "cred" not in cols:
                    cols["cred"] = i
                elif "oficina" in v and "oficina" not in cols:
                    cols["oficina"] = i
                elif ("nformaci" in v or "adicional" in v) and "info" not in cols:
                    cols["info"] = i
            return r, cols
    raise ValueError(
        "Banco: no se encontró la fila de encabezado con 'Fecha' y 'Débito'. "
        "Verificá que el .xlsx descargado es el extracto correcto de Caja Social."
    )


# ── BANCO RAW (para Hoja1, filtrado al rango pedido) ──────────────────────────
# El nuevo extracto trae todo el histórico (3+ meses), así que sí o sí hay que
# filtrar por fecha aquí. Si no, Hoja1 traería movs ajenos al mes conciliado y
# las sumas de impuestos al pie de CREDITO quedarían infladas.
def leer_banco_raw(data: bytes, fecha_ini: date, fecha_fin: date):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[BANCO_SHEET] if BANCO_SHEET in wb.sheetnames else wb.active

    header_row, cols = _localizar_header_banco(ws)
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        fecha_v  = ws.cell(r, cols["fecha"]).value      if "fecha"   in cols else None
        desc_v   = ws.cell(r, cols["desc"]).value       if "desc"    in cols else None
        doc_v    = ws.cell(r, cols["doc"]).value        if "doc"     in cols else None
        deb_v    = ws.cell(r, cols["deb"]).value        if "deb"     in cols else None
        cred_v   = ws.cell(r, cols["cred"]).value       if "cred"    in cols else None
        ofic_v   = ws.cell(r, cols["oficina"]).value    if "oficina" in cols else None
        info_v   = ws.cell(r, cols["info"]).value       if "info"    in cols else None

        if fecha_v is None and desc_v is None and deb_v in (None, "") and cred_v in (None, ""):
            continue

        desc = str(desc_v).strip() if desc_v is not None else ""
        if desc in ("SALDO INICIAL", "SALDO FINAL"):
            continue

        f = parse_fecha(fecha_v)
        if not f or f < fecha_ini or f > fecha_fin:
            continue

        deb_n  = parse_monto_co(deb_v)
        cred_n = parse_monto_co(cred_v)
        if deb_n == 0 and cred_n == 0 and not desc:
            continue

        # Valor con signo: Crédito positivo, Débito negativo
        valor = cred_n - deb_n
        valor = int(valor) if valor == int(valor) else valor

        # Filas con descripciones especiales (-- vacíos)
        info_str = ""
        if info_v is not None and str(info_v).strip() not in ("", "--"):
            info_str = str(info_v).strip()

        row_out = [
            f.strftime("%d/%m/%Y"),                                     # Fecha Transacción
            desc,                                                       # Descripción
            valor if valor != 0 else 0,                                 # Valor (signed)
            "",                                                         # Saldo (no disponible)
            "",                                                         # Regional / Oficina (no disponible)
            "",                                                         # Tipo Transacción (no disponible)
            str(ofic_v).strip() if ofic_v is not None else "",          # Oficina
            str(doc_v).strip() if doc_v is not None else "",            # Ref. Titular Cuenta (← Documento)
            info_str,                                                   # Detalles Adicionales
        ]
        rows.append(row_out)
    return list(HOJA1_HEADERS), rows


# ── BANCO FILTRADO (para DEBITO / CREDITO) ────────────────────────────────────
def leer_banco(data: bytes, fecha_ini: date, fecha_fin: date):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[BANCO_SHEET] if BANCO_SHEET in wb.sheetnames else wb.active

    header_row, cols = _localizar_header_banco(ws)
    positivos, negativos = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        fecha_v = ws.cell(r, cols["fecha"]).value if "fecha" in cols else None
        desc_v  = ws.cell(r, cols["desc"]).value  if "desc"  in cols else None
        deb_v   = ws.cell(r, cols["deb"]).value   if "deb"   in cols else None
        cred_v  = ws.cell(r, cols["cred"]).value  if "cred"  in cols else None

        desc = str(desc_v).strip() if desc_v is not None else ""
        if desc in ("SALDO INICIAL", "SALDO FINAL"):
            continue
        f = parse_fecha(fecha_v)
        if not f or f < fecha_ini or f > fecha_fin:
            continue

        deb_n  = parse_monto_co(deb_v)
        cred_n = parse_monto_co(cred_v)
        valor  = cred_n - deb_n
        if valor == 0:
            continue

        fecha_str = f.strftime("%d/%m/%Y")
        if valor > 0:
            positivos.append((fecha_str, desc, valor))
        elif desc not in IMPTOS_SET:
            # Antes filtrábamos por "tipo" (N005/N328/...). Ahora el extracto
            # ya no trae esa columna, así que excluimos por descripción.
            negativos.append((fecha_str, desc, abs(valor), ""))
    return positivos, negativos


# ── SIIGO ─────────────────────────────────────────────────────────────────────
def leer_siigo(data: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 3).value
        if val and str(val).strip() == "Comprobante":
            header_row = r
            break
    if header_row is None:
        raise ValueError("No se encontró la cabecera del reporte Siigo (columna 'Comprobante' en col C)")

    debitos, creditos = [], []
    for r in range(header_row + 2, ws.max_row + 1):
        comp      = ws.cell(r, 3).value
        fecha_raw = ws.cell(r, 5).value
        deb_val   = ws.cell(r, 13).value
        cred_val  = ws.cell(r, 14).value

        if comp is None and deb_val is None and cred_val is None: continue
        comp_str = str(comp).strip() if comp else ""
        if not comp_str or comp_str.startswith("Total") or comp_str.startswith("Cuenta contable"):
            continue

        if fecha_raw is None:
            fecha_str = ""
        elif hasattr(fecha_raw, "strftime"):
            fecha_str = fecha_raw.strftime("%d/%m/%Y")
        else:
            fecha_str = str(fecha_raw).strip()

        deb  = float(deb_val)  if deb_val  else 0.0
        cred = float(cred_val) if cred_val else 0.0

        if deb  > 0: debitos.append((comp_str, fecha_str, deb))
        if cred > 0:
            es_cc10 = comp_str.upper().startswith("CC-10")
            creditos.append((comp_str, fecha_str, cred, es_cc10))

    return debitos, creditos


# ── GENERAR EXCEL ─────────────────────────────────────────────────────────────
def generar_excel(banco_raw_headers, banco_raw_rows,
                  positivos_banco, negativos_banco,
                  siigo_deb, siigo_cred,
                  fecha_ini: date, fecha_fin: date):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    FONT  = "Trebuchet MS"
    SZ    = 12
    FMT   = '#,##0.00'

    # Paleta pastel — mismas tonalidades, más suaves
    C_GREEN   = "FFE8F5E9"   # menta pastel
    C_YELLOW  = "FFFFFCCC"   # crema-limón
    C_RED     = "FFFFEBEE"   # rosa pastel
    C_HEADER  = "FF4472C4"   # azul encabezado (igual que DIAN)
    C_GRAY    = "FFD3D3D3"   # encabezado Hoja1

    def font(bold=False, color="FF000000"):
        return Font(name=FONT, size=SZ, bold=bold, color=color)

    def fill(hex8):
        return PatternFill("solid", fgColor=hex8)

    # Borde default de Excel (thin negro) — igual para todas las celdas de CREDITO
    _side = Side(style="thin", color="FF000000")
    CRED_BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)

    def center():
        return Alignment(horizontal="center")

    wb = openpyxl.Workbook()

    # ══════════════════════════════════════════════════════════
    # HOJA 1 — extracto bancario completo (9 columnas)
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Hoja1"

    for c, h in enumerate(banco_raw_headers, 1):
        cell = ws1.cell(1, c, h)
        cell.font      = font(bold=True)
        cell.fill      = fill(C_GRAY)
        cell.alignment = center()

    ws1.auto_filter.ref = f"A1:{ws1.cell(1, len(banco_raw_headers)).column_letter}1"

    for i, w in enumerate([14, 38, 14, 16, 14, 14, 28, 20, 40], 1):
        ws1.column_dimensions[ws1.cell(1, i).column_letter].width = w

    for r, row_data in enumerate(banco_raw_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(r, c, val if val != "" else None)
            cell.font = font()
            if c == 3: cell.number_format = FMT
            if c == 4: cell.number_format = FMT

    # ══════════════════════════════════════════════════════════
    # DEBITO — A-C Siigo | D sep | E Fecha | F Banco | G Diferencia | H Siigo
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("DEBITO")

    for c, h in enumerate(["Comprobante", "Fecha elaboración", "Débito"], 1):
        cell = ws2.cell(1, c, h)
        cell.font = font(bold=True, color="FFFFFFFF")
        cell.fill = fill(C_HEADER)
        cell.alignment = center()

    for c, h in enumerate(["Fecha", "Banco (suma día)", "Diferencia", "Siigo (suma día)"], 5):
        cell = ws2.cell(1, c, h)
        cell.font = font(bold=True, color="FFFFFFFF")
        cell.fill = fill(C_HEADER)
        cell.alignment = center()

    for col, w in {"A":16,"B":18,"C":16,"D":8,"E":14,"F":18,"G":16,"H":18}.items():
        ws2.column_dimensions[col].width = w

    siigo_deb_sorted = sorted(siigo_deb, key=lambda x: parse_fecha(x[1]) or date.min)
    for r, (comp, fecha_str, monto) in enumerate(siigo_deb_sorted, 2):
        ws2.cell(r, 1, comp).font = font()
        ws2.cell(r, 2, fecha_str).font = font()
        c3 = ws2.cell(r, 3, monto); c3.font = font(); c3.number_format = FMT

    banco_por_fecha = defaultdict(float)
    for fs, _, val in positivos_banco:
        f = parse_fecha(fs)
        if f: banco_por_fecha[f] += val

    siigo_por_fecha = defaultdict(float)
    for _, fs, monto in siigo_deb:
        f = parse_fecha(fs)
        if f: siigo_por_fecha[f] += monto

    all_dates = sorted(set(list(banco_por_fecha) + list(siigo_por_fecha)))
    for r, f in enumerate(all_dates, 2):
        banco_sum = banco_por_fecha.get(f, 0.0)
        siigo_sum = siigo_por_fecha.get(f, 0.0)
        diff      = banco_sum - siigo_sum

        ws2.cell(r, 5, f.strftime("%d/%m/%Y")).font = font()
        c6 = ws2.cell(r, 6, banco_sum); c6.font = font(); c6.number_format = FMT

        diff_val   = diff if abs(diff) > 0.01 else 0.0
        diff_color = "FFFF0000" if abs(diff) > 0.01 else "FF000000"
        c7 = ws2.cell(r, 7, diff_val); c7.font = font(color=diff_color); c7.number_format = FMT

        c8 = ws2.cell(r, 8, siigo_sum); c8.font = font(); c8.number_format = FMT

    if all_dates:
        total_diff = round(sum(banco_por_fecha.get(f,0)-siigo_por_fecha.get(f,0) for f in all_dates), 2)
        ct = ws2.cell(len(all_dates) + 2, 7, total_diff)
        ct.font = font(bold=True); ct.number_format = FMT

    # ══════════════════════════════════════════════════════════
    # CREDITO — Verde → Amarillo → Rojo, por valor desc dentro de cada grupo
    # Al final: fila en blanco + 5 filas de impuestos (suma de Hoja1)
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("CREDITO")

    for c, h in enumerate(["Fecha Transacción","Descripción","Valor",
                            "Diferencia","Crédito Siigo","Comprobante","Fecha Siigo"], 1):
        cell = ws3.cell(1, c, h)
        cell.font = font(bold=True, color="FFFFFFFF")
        cell.fill = fill(C_HEADER)
        cell.alignment = center()

    for col, w in {"A":18,"B":38,"C":16,"D":16,"E":16,"F":16,"G":18}.items():
        ws3.column_dimensions[col].width = w

    # Pool de negativos banco
    banco_pool = defaultdict(list)
    for fecha_str, desc, val_abs, tipo in negativos_banco:
        banco_pool[round(val_abs, 2)].append((fecha_str, desc))

    # Clasificar créditos Siigo
    siigo_all = sorted(siigo_cred, key=lambda x: -x[2])
    verde_rows, amarillo_rows = [], []

    for comp, siigo_fecha, siigo_val, es_cc10 in siigo_all:
        key = round(siigo_val, 2)
        if banco_pool.get(key):
            banco_fecha, banco_desc = banco_pool[key].pop(0)
            rd = dict(sort_val=siigo_val, tipo="cc10_match" if es_cc10 else "match",
                      banco_fecha=banco_fecha, banco_desc=banco_desc,
                      banco_val=-siigo_val, siigo_val=siigo_val,
                      comp=comp, siigo_fecha=siigo_fecha)
            (amarillo_rows if es_cc10 else verde_rows).append(rd)
        else:
            amarillo_rows.append(dict(sort_val=siigo_val, tipo="siigo_only",
                                      banco_fecha="", banco_desc="",
                                      banco_val=0, siigo_val=siigo_val,
                                      comp=comp, siigo_fecha=siigo_fecha))

    rojo_rows = []
    for key, entries in banco_pool.items():
        for fecha_str, desc in entries:
            rojo_rows.append(dict(sort_val=key, tipo="banco_only",
                                  banco_fecha=fecha_str, banco_desc=desc))

    verde_rows.sort(key=lambda x: -x["sort_val"])
    amarillo_rows.sort(key=lambda x: -x["sort_val"])
    rojo_rows.sort(key=lambda x: -x["sort_val"])

    all_cred_rows = (
        [("siigo", r) for r in verde_rows] +
        [("siigo", r) for r in amarillo_rows] +
        [("banco", r) for r in rojo_rows]
    )

    for row_num, (row_type, rd) in enumerate(all_cred_rows, 2):
        tipo = rd["tipo"]
        if tipo == "match":
            fc = C_GREEN
        elif tipo in ("cc10_match", "siigo_only"):
            fc = C_YELLOW
        else:
            fc = C_RED

        fl = fill(fc)

        def wc(col, val, fmt=None):
            cell = ws3.cell(row_num, col, val)
            cell.font = font(); cell.fill = fl; cell.border = CRED_BORDER
            if fmt: cell.number_format = fmt
            return cell

        if tipo in ("match", "cc10_match"):
            wc(1, rd["banco_fecha"]); wc(2, rd["banco_desc"])
            wc(3, rd["banco_val"], FMT)
            diff = rd["banco_val"] + rd["siigo_val"]
            wc(4, diff if abs(diff) > 0.01 else None, FMT)
            wc(5, rd["siigo_val"], FMT)
            wc(6, rd["comp"]); wc(7, rd["siigo_fecha"])

        elif tipo == "siigo_only":
            for c in range(1, 5):
                cell = ws3.cell(row_num, c); cell.fill = fl; cell.border = CRED_BORDER
            wc(5, rd["siigo_val"], FMT); wc(6, rd["comp"]); wc(7, rd["siigo_fecha"])

        else:  # banco_only
            wc(1, rd["banco_fecha"]); wc(2, rd["banco_desc"])
            wc(3, -rd["sort_val"], FMT)
            for c in range(4, 8):
                cell = ws3.cell(row_num, c); cell.fill = fl; cell.border = CRED_BORDER

    # ── Sección de impuestos al final de CREDITO ──────────────────────────────
    # Suma de Hoja1 col Valor (idx 2) agrupada por descripción (idx 1)
    imptos_sums = {}
    for desc in IMPTOS_DESCS:
        total = sum(row[2] for row in banco_raw_rows
                    if row[1] == desc and isinstance(row[2], (int, float)))
        imptos_sums[desc] = round(total, 2)

    last_data_row = len(all_cred_rows) + 1   # +1 por encabezado
    imptos_start  = last_data_row + 2         # +1 fila en blanco + 1

    for i, desc in enumerate(IMPTOS_DESCS):
        r = imptos_start + i
        val = imptos_sums[desc]

        ws3.cell(r, 1).border = CRED_BORDER

        cb = ws3.cell(r, 2, desc)
        cb.font = font(); cb.border = CRED_BORDER

        cc = ws3.cell(r, 3, val)
        cc.font = font(); cc.number_format = FMT; cc.border = CRED_BORDER

        for c in range(4, 8):
            ws3.cell(r, c).border = CRED_BORDER

    # ── Nombre de archivo ─────────────────────────────────────────────────────
    meses = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
             7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
    filename = f"CTA AHORROS {meses.get(fecha_ini.month, str(fecha_ini.month))} {fecha_ini.year}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), filename


# ── VALIDACIÓN (antes de procesar) ───────────────────────────────────────────
def _norm_s(s):
    """Minúsculas sin tildes para comparación flexible."""
    if s is None:
        return ''
    n = unicodedata.normalize('NFKD', str(s)).encode('ascii', 'ignore').decode()
    return n.lower().strip()


def _validar_banco_xlsx(banco_bytes):
    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(banco_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f'Banco: no se pudo abrir el archivo .xlsx ({e}).')

    if BANCO_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Banco: la hoja '{BANCO_SHEET}' no está en el archivo "
            f"(hojas encontradas: {wb.sheetnames}). "
            f"Verificá que es el extracto descargado del portal de Caja Social."
        )
    ws = wb[BANCO_SHEET]
    # _localizar_header_banco ya valida la presencia de Fecha y Débito y lanza
    # ValueError descriptivo si no los encuentra.
    _localizar_header_banco(ws)


def _validar_siigo_cta(siigo_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(siigo_bytes), data_only=True)
    ws = wb.active

    header_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 3).value
        if val and str(val).strip() == 'Comprobante':
            header_row = r
            break
    if header_row is None:
        raise ValueError(
            "Siigo: No se encontró 'Comprobante' en columna C. "
            "Verificá que el archivo es el reporte correcto exportado desde Siigo."
        )

    num_cols = ws.max_column
    if num_cols < 14:
        raise ValueError(
            f'Siigo: Se esperan al menos 14 columnas, el archivo tiene {num_cols}. '
            f'Verificá que es el reporte de Siigo completo.'
        )

    h13 = ws.cell(header_row, 13).value  # columna M — debe ser Débito
    h14 = ws.cell(header_row, 14).value  # columna N — debe ser Crédito
    errores = []
    if 'deb' not in _norm_s(h13):
        errores.append(f'col M (13) tiene "{h13}", se esperaba "Débito"')
    if 'cred' not in _norm_s(h14):
        errores.append(f'col N (14) tiene "{h14}", se esperaba "Crédito"')

    if errores:
        disponibles = [ws.cell(header_row, c).value for c in range(1, num_cols + 1)]
        raise ValueError(
            f'Siigo: Columnas de montos inesperadas — {"; ".join(errores)}. '
            f'Encabezados disponibles: {disponibles}'
        )


# ── HANDLER VERCEL ────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            body   = json.loads(self.rfile.read(length))

            f1 = body.get("banco_b64", "") or body.get("file1_b64", "")
            f2 = body.get("siigo_b64", "") or body.get("file2_b64", "")
            if not f1 or not f2:
                raise ValueError("Faltan archivos requeridos")

            banco_bytes, siigo_bytes = detectar_archivos(
                base64.b64decode(f1), base64.b64decode(f2)
            )

            _validar_banco_xlsx(banco_bytes)
            _validar_siigo_cta(siigo_bytes)

            fecha_ini = datetime.strptime(body["fecha_inicio"], "%Y-%m-%d").date()
            fecha_fin = datetime.strptime(body["fecha_fin"],    "%Y-%m-%d").date()

            banco_raw_headers, banco_raw_rows = leer_banco_raw(banco_bytes, fecha_ini, fecha_fin)
            positivos, negativos              = leer_banco(banco_bytes, fecha_ini, fecha_fin)
            siigo_deb, siigo_cred             = leer_siigo(siigo_bytes)

            excel_bytes, filename = generar_excel(
                banco_raw_headers, banco_raw_rows,
                positivos, negativos,
                siigo_deb, siigo_cred,
                fecha_ini, fecha_fin
            )

            total_deb_banco  = round(sum(r[2] for r in banco_raw_rows if isinstance(r[2], (int,float)) and r[2] > 0), 2)
            total_cred_banco = round(sum(abs(r[2]) for r in banco_raw_rows if isinstance(r[2], (int,float)) and r[2] < 0), 2)
            total_deb_siigo  = round(sum(m for _,_,m in siigo_deb), 2)
            total_cred_siigo = round(sum(m for _,_,m,_ in siigo_cred), 2)
            diff_deb         = round(total_deb_banco  - total_deb_siigo,  2)
            diff_cred        = round(total_cred_banco - total_cred_siigo, 2)

            response = {
                "archivo_b64": base64.b64encode(excel_bytes).decode(),
                "nombre":      filename,
                "resumen": {
                    "banco_positivos":   len(positivos),
                    "banco_negativos":   len(negativos),
                    "siigo_debitos":     len(siigo_deb),
                    "siigo_creditos":    len(siigo_cred),
                    "total_deb_banco":   total_deb_banco,
                    "total_deb_siigo":   total_deb_siigo,
                    "diff_deb":          diff_deb,
                    "total_cred_banco":  total_cred_banco,
                    "total_cred_siigo":  total_cred_siigo,
                    "diff_cred":         diff_cred,
                    "conciliado":        abs(diff_deb) < 1 and abs(diff_cred) < 1,
                }
            }

            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps(response).encode())

        except Exception as e:
            self.send_response(500)
            self.send_header("Content-Type", "application/json")
            self.end_headers()
            self.wfile.write(json.dumps({"error": str(e)}).encode())

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
