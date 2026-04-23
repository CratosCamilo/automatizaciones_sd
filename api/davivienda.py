"""
Vercel Python serverless function — Conciliación Davivienda + Redeban QR.

Recibe:
- Excel (.xlsx) de Davivienda sin arreglar
- CSV de Redeban QR (consulta de transacciones)
- Rango de fechas (fecha_inicio, fecha_fin)

Devuelve: Excel con un día por fila — columna 'Descripción motivo' ya tiene los
nombres de las personas que consignaron (tomados del CSV de Redeban para las
filas que Davivienda reporta como 'Pago A Llave De Comercio').
"""

from http.server import BaseHTTPRequestHandler
import base64
import csv
import io
import json
from collections import defaultdict
from datetime import datetime, date

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


FUENTE    = 'Trebuchet MS'
SIZE_BODY = 12
FMT_NUM   = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'
RED       = 'FFFF0000'
HEADER_BG = 'FFB2AEAE'

LLAVE_KEY = 'Pago A Llave De Comercio'


# ─────────────────────────────────────────────────────────────────────────────
# PARSERS
# ─────────────────────────────────────────────────────────────────────────────

def _parse_fecha_dd(s):
    """'01/04/2026' (o datetime/date) → date."""
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    dd, mm, yyyy = str(s).strip().split('/')
    return date(int(yyyy), int(mm), int(dd))


def _parse_valor_dav(v):
    """'$ 400.000,00' → 400000 (int si entero) o 400000.5 (float con 2 decimales)."""
    if v is None:
        return None
    s = str(v).replace('$', '').replace(' ', '').replace('\xa0', '').strip()
    if not s or s in ('-', '--'):
        return None
    # Formato español: '.' miles, ',' decimal → quitar puntos, cambiar coma por punto
    s = s.replace('.', '').replace(',', '.')
    try:
        f = float(s)
    except ValueError:
        return None
    return int(f) if f.is_integer() else round(f, 2)


def _match_key(fecha, valor):
    """Normaliza (fecha, valor) a int centavos para comparación exacta."""
    return (fecha, int(round(float(valor) * 100)))


def parse_davivienda(wb, fecha_ini, fecha_fin):
    ws = wb.active
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(1, c).value
        if val is None:
            continue
        headers[str(val).strip().lower()] = c

    col_fecha = headers.get('fecha de sistema')
    col_desc  = headers.get('descripción motivo') or headers.get('descripcion motivo')
    col_trans = headers.get('transacción') or headers.get('transaccion')
    col_valor = headers.get('valor total')
    if not all([col_fecha, col_desc, col_trans, col_valor]):
        raise ValueError(
            f'No se encontraron columnas esperadas en Davivienda. '
            f'Encabezados leídos: {list(headers.keys())}'
        )

    rows = []
    for r in range(2, ws.max_row + 1):
        fecha_raw = ws.cell(r, col_fecha).value
        if not fecha_raw:
            continue
        try:
            fecha = _parse_fecha_dd(fecha_raw)
        except Exception:
            continue
        if fecha < fecha_ini or fecha > fecha_fin:
            continue

        desc  = ws.cell(r, col_desc).value or ''
        trans = (ws.cell(r, col_trans).value or '').strip()
        valor = _parse_valor_dav(ws.cell(r, col_valor).value)
        if valor is None:
            continue

        rows.append({
            'fecha':     fecha,
            'fecha_str': str(fecha_raw).strip(),
            'desc':      str(desc),
            'trans':     trans,
            'valor':     valor,
        })
    return rows


def parse_redeban(csv_bytes, fecha_ini, fecha_fin):
    # Probar utf-8-sig primero (Redeban exporta con BOM a veces), luego latin-1
    try:
        text = csv_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = csv_bytes.decode('latin-1')

    reader = csv.reader(io.StringIO(text), delimiter=';')
    headers = next(reader, None)
    if not headers:
        return []

    def find(name):
        target = name.lower()
        for i, h in enumerate(headers):
            if (h or '').strip().lower() == target:
                return i
        return None

    idx_estado = find('estado')
    idx_emisor = find('emisor')
    idx_valor  = find('valor')
    idx_fecha  = find('fecha')
    if None in (idx_estado, idx_emisor, idx_valor, idx_fecha):
        raise ValueError(
            f'CSV de Redeban sin columnas esperadas. Encabezados: {headers}'
        )

    entries = []
    for row in reader:
        if len(row) <= max(idx_estado, idx_emisor, idx_valor, idx_fecha):
            continue
        if row[idx_estado].strip().upper() != 'ACEPTADA':
            continue
        try:
            v = float(row[idx_valor])
        except ValueError:
            continue
        valor = int(v) if v.is_integer() else round(v, 2)
        raw_fecha = row[idx_fecha].strip()
        try:
            fecha = datetime.strptime(raw_fecha[:10], '%Y-%m-%d').date()
        except ValueError:
            continue
        if fecha < fecha_ini or fecha > fecha_fin:
            continue
        entries.append({
            'fecha':  fecha,
            'valor':  valor,
            'emisor': row[idx_emisor].strip(),
        })
    return entries


# ─────────────────────────────────────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────────────────────────────────────

def aplicar_nombres(dav_rows, redeban_entries):
    """En las filas 'Pago A Llave De Comercio' de Davivienda, reemplaza la
    descripción con el Emisor del CSV que matchee por (fecha, valor)."""
    bucket = defaultdict(list)
    for e in redeban_entries:
        bucket[_match_key(e['fecha'], e['valor'])].append(e['emisor'])

    matched = unmatched = 0
    for r in dav_rows:
        if r['desc'].strip() != LLAVE_KEY:
            continue
        key = _match_key(r['fecha'], r['valor'])
        if bucket[key]:
            r['desc'] = bucket[key].pop(0)
            matched += 1
        else:
            unmatched += 1

    extras = sum(len(v) for v in bucket.values())
    return matched, unmatched, extras


# ─────────────────────────────────────────────────────────────────────────────
# ESCRITURA XLSX
# ─────────────────────────────────────────────────────────────────────────────

def generar_excel(rows):
    # Orden final: Fecha ASC, Valor ASC. Desempate: Descripción DESC (Z→A), que
    # hace que nombres de personas (C***, U***) vayan antes de 'Abono ACH…'
    # cuando coinciden fecha y valor. Se aplica con dos sorts estables.
    rows.sort(key=lambda r: r['desc'], reverse=True)
    rows.sort(key=lambda r: (r['fecha'], r['valor']))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Movimientos1'

    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    header_font = Font(name=FUENTE, size=SIZE_BODY, bold=True)

    for c, name in enumerate(['Fecha de Sistema', 'Descripción motivo', 'Transacción', 'Valor Total'], start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
    ws.cell(row=1, column=4).number_format = FMT_NUM

    body_font = Font(name=FUENTE, size=SIZE_BODY)
    red_font  = Font(name=FUENTE, size=SIZE_BODY, color=RED)

    for i, r in enumerate(rows, start=2):
        font = red_font if r['trans'] == 'Nota Débito' else body_font

        c1 = ws.cell(row=i, column=1, value=r['fecha_str'])
        c2 = ws.cell(row=i, column=2, value=r['desc'])
        c3 = ws.cell(row=i, column=3, value=r['trans'])
        c4 = ws.cell(row=i, column=4, value=r['valor'])
        for c in (c1, c2, c3, c4):
            c.font = font
        c4.number_format = FMT_NUM

    ws.column_dimensions['A'].width = 13.28515625
    ws.column_dimensions['B'].width = 73.42578125
    ws.column_dimensions['C'].width = 15.5703125
    ws.column_dimensions['D'].width = 20.140625

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# HANDLER HTTP
# ─────────────────────────────────────────────────────────────────────────────

MESES = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',
         7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}


class handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin',  '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
            data   = json.loads(body)

            for field in ('davivienda', 'redeban', 'fecha_inicio', 'fecha_fin'):
                if field not in data:
                    return self._error(400, f'Falta el campo "{field}"')

            fecha_ini = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(data['fecha_fin'],    '%Y-%m-%d').date()

            dav_bytes = base64.b64decode(data['davivienda'])
            csv_bytes = base64.b64decode(data['redeban'])

            wb = openpyxl.load_workbook(io.BytesIO(dav_bytes))
            dav_rows = parse_davivienda(wb, fecha_ini, fecha_fin)
            redeban_entries = parse_redeban(csv_bytes, fecha_ini, fecha_fin)

            matched, unmatched, extras = aplicar_nombres(dav_rows, redeban_entries)
            output = generar_excel(dav_rows)

            mes = MESES.get(fecha_ini.month, str(fecha_ini.month))
            filename = f'SEMANA {fecha_ini.day} AL {fecha_fin.day} {mes} {fecha_ini.year}.xlsx'

            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('X-Total-Rows', str(len(dav_rows)))
            self.send_header('X-Matched',    str(matched))
            self.send_header('X-Unmatched',  str(unmatched))
            self.send_header('X-Extras',     str(extras))
            self._cors_headers()
            self.send_header('Access-Control-Expose-Headers',
                'X-Total-Rows, X-Matched, X-Unmatched, X-Extras')
            self.end_headers()
            self.wfile.write(output)

        except json.JSONDecodeError:
            self._error(400, 'El cuerpo de la solicitud no es JSON válido')
        except Exception as e:
            self._error(500, str(e))

    def _error(self, code, message):
        body = json.dumps({'error': message}).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self._cors_headers()
        self.end_headers()
        self.wfile.write(body)
