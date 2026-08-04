"""
Vercel Python serverless function — Inventario mensual Zapatoca.

Recibe 3 excels (Siigo pólizas detalladas, INVENTARIO YYYY, Stock al fecha) en
base64 vía JSON POST. El orden NO importa: el archivo se detecta por contenido.

Devuelve el mismo INVENTARIO YYYY con una hoja nueva del mes detectado del stock,
duplicada de la hoja del mes anterior, con:
  - INV INICIAL ← INV FINAL del mes anterior (valores calculados)
  - UNITARIO conservado
  - COMPRAS rellenada desde el pool de Siigo (matcheando por nombre)
  - CANTIDAD rellenada desde el stock físico (matcheando por nombre)
  - INV FINAL / COSTEO como fórmulas
  - Bloque de sobrantes a la derecha (col I-K) con nombres que no matchearon
"""

from http.server import BaseHTTPRequestHandler
import base64
import io
import json
import re

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

MESES = [
    'ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO',
    'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE',
]

MESES_SET = set(MESES)

# Columnas de la hoja INVENTARIO (fila 2 headers, datos desde fila 3)
COL_NOMBRE      = 1  # A
COL_INV_INICIAL = 2  # B
COL_COMPRAS     = 3  # C
COL_UNITARIO    = 4  # D
COL_CANTIDAD    = 5  # E
COL_INV_FINAL   = 6  # F
COL_COSTEO      = 7  # G

# Bloque de sobrantes: col H libre como separador, bloque en I-K
COL_SOB_NOMBRE   = 9   # I
COL_SOB_COMPRAS  = 10  # J
COL_SOB_CANTIDAD = 11  # K

FILTRO_DETALLE = 'PRODUCTO: MATERIA PRIMA VARIOS'  # busca case-insensitive

# Pólizas: layout observado
POL_HEADER_ROW    = 8
POL_COL_DESCRIP   = 8   # H
POL_COL_DETALLE   = 9   # I
POL_COL_DEBITO    = 11  # K

# Formatos visuales para el bloque de sobrantes
FUENTE = 'Trebuchet MS'
COLOR_HEADER_SOB_BG = 'FDE9D9'  # naranja claro (mismo que "extra" en zapatoca)
COLOR_HEADER_SOB_FG = '000000'


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def _s(v):
    """String o None."""
    if v is None:
        return ''
    return str(v)


def _num(v):
    """Convierte a float o 0."""
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _sheet_names_normalized(wb):
    """Nombres de hojas normalizados (upper, sin tildes básicas, strip)."""
    out = []
    for sn in wb.sheetnames:
        n = sn.strip().upper()
        # openpyxl a veces devuelve nombres con encoding raro; solo mayúsculas
        out.append(n)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# DETECCIÓN DE ARCHIVOS
# ─────────────────────────────────────────────────────────────────────────────

def _es_inventario(wb):
    """True si el workbook parece un INVENTARIO YYYY.
    Regla: tiene al menos una hoja con nombre de mes, y la fila 2 de esa hoja
    tiene los headers esperados.
    """
    for sn in wb.sheetnames:
        if sn.strip().upper() not in MESES_SET:
            continue
        ws = wb[sn]
        h_nombre = _s(ws.cell(2, COL_NOMBRE).value).strip().upper()
        h_inv    = _s(ws.cell(2, COL_INV_INICIAL).value).strip().upper()
        h_comp   = _s(ws.cell(2, COL_COMPRAS).value).strip().upper()
        if h_nombre == 'NOMBRE' and 'INV INICIAL' in h_inv and h_comp == 'COMPRAS':
            return True
    return False


def _es_polizas(wb):
    """True si el workbook parece pólizas detalladas de Siigo.
    Regla: alguna hoja tiene "Detalle" en I8 y al menos una fila con
    "Producto: MATERIA PRIMA VARIOS" en col I después de la fila 8.
    """
    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < POL_HEADER_ROW:
            continue
        h_detalle = _s(ws.cell(POL_HEADER_ROW, POL_COL_DETALLE).value).strip().upper()
        if 'DETALLE' not in h_detalle:
            continue
        # buscar al menos una fila con el patrón
        for r in range(POL_HEADER_ROW + 1, min(ws.max_row, POL_HEADER_ROW + 200) + 1):
            det = _s(ws.cell(r, POL_COL_DETALLE).value).upper()
            if FILTRO_DETALLE in det:
                return True
    return False


def _es_stock(wb):
    """True si el workbook parece stock físico.
    Regla: alguna hoja tiene B1 que empieza con "Stock al " seguido de fecha.
    """
    for sn in wb.sheetnames:
        ws = wb[sn]
        b1 = _s(ws.cell(1, 2).value).strip()
        if b1.upper().startswith('STOCK AL'):
            # además col A1 debe ser algo tipo "Producto"
            a1 = _s(ws.cell(1, 1).value).strip().upper()
            if a1.startswith('PRODUCTO'):
                return True
    return False


def detectar_archivos(cargados):
    """cargados: lista de tuplas (nombre_original, workbook_bytes).
    Retorna dict {'inventario': bytes, 'polizas': bytes, 'stock': bytes}.
    """
    resultado = {'inventario': None, 'polizas': None, 'stock': None}

    for _, bytes_ in cargados:
        # Cargamos con data_only=False para leer estructura/nombres.
        wb = openpyxl.load_workbook(io.BytesIO(bytes_), data_only=True)
        es_inv = _es_inventario(wb)
        es_pol = _es_polizas(wb)
        es_stk = _es_stock(wb)

        # Solo debería matchear una regla
        matches = [k for k, v in {'inventario': es_inv, 'polizas': es_pol, 'stock': es_stk}.items() if v]
        if len(matches) == 0:
            raise ValueError(
                'Uno de los archivos subidos no coincide con ninguno de los 3 esperados '
                '(INVENTARIO, Pólizas detalladas Siigo, Stock al YYYY-MM-DD). '
                'Verifica que subiste los archivos correctos.'
            )
        if len(matches) > 1:
            raise ValueError(
                f'Un archivo subido matchea ambiguamente como {matches}. '
                'Revisa que los archivos no estén mezclados entre sí.'
            )

        tipo = matches[0]
        if resultado[tipo] is not None:
            raise ValueError(
                f'Se subieron dos archivos del mismo tipo ({tipo}). '
                'Necesito exactamente uno de cada: INVENTARIO, Pólizas, Stock.'
            )
        resultado[tipo] = bytes_

    faltantes = [k for k, v in resultado.items() if v is None]
    if faltantes:
        raise ValueError(f'Faltan archivos: {faltantes}')

    return resultado


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN DEL MES OBJETIVO (desde stock)
# ─────────────────────────────────────────────────────────────────────────────

def extraer_mes_desde_stock(stock_bytes):
    """Del header B1 tipo "Stock al 2026-07-31" retorna (mes_nombre, año, día)."""
    wb = openpyxl.load_workbook(io.BytesIO(stock_bytes), data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        b1 = _s(ws.cell(1, 2).value).strip()
        m = re.search(r'(\d{4})-(\d{2})-(\d{2})', b1)
        if m:
            anio  = int(m.group(1))
            mes_n = int(m.group(2))
            dia   = int(m.group(3))
            if 1 <= mes_n <= 12:
                return MESES[mes_n - 1], anio, dia
    raise ValueError(
        'No pude leer la fecha del archivo de stock. Esperaba un header tipo '
        '"Stock al YYYY-MM-DD" en la celda B1.'
    )


# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUIR POOL DESDE PÓLIZAS
# ─────────────────────────────────────────────────────────────────────────────

def construir_pool(polizas_bytes):
    """Retorna dict {nombre.strip(): suma_debito}."""
    wb = openpyxl.load_workbook(io.BytesIO(polizas_bytes), data_only=True)
    # Buscar la hoja correcta (la que tiene "Detalle" en I8)
    ws = None
    for sn in wb.sheetnames:
        cand = wb[sn]
        if cand.max_row < POL_HEADER_ROW:
            continue
        if 'DETALLE' in _s(cand.cell(POL_HEADER_ROW, POL_COL_DETALLE).value).strip().upper():
            ws = cand
            break
    if ws is None:
        raise ValueError('No encontré una hoja con encabezado "Detalle" en Pólizas.')

    pool = {}
    for r in range(POL_HEADER_ROW + 1, ws.max_row + 1):
        detalle = _s(ws.cell(r, POL_COL_DETALLE).value).upper()
        if FILTRO_DETALLE not in detalle:
            continue
        nombre = _s(ws.cell(r, POL_COL_DESCRIP).value).strip()
        debito = _num(ws.cell(r, POL_COL_DEBITO).value)
        if not nombre or debito == 0:
            continue
        pool[nombre] = pool.get(nombre, 0.0) + debito
    return pool


# ─────────────────────────────────────────────────────────────────────────────
# CONSTRUIR DICCIONARIO DE STOCK
# ─────────────────────────────────────────────────────────────────────────────

def construir_stock(stock_bytes):
    """Retorna dict {nombre.strip(): cantidad} solo de la hoja Producción."""
    wb = openpyxl.load_workbook(io.BytesIO(stock_bytes), data_only=True)
    ws = None
    for sn in wb.sheetnames:
        # buscar la hoja "Producción" (con o sin tilde), fallback a la primera con schema válido
        if 'PRODUCC' in sn.upper():
            ws = wb[sn]
            break
    if ws is None:
        # fallback a la activa si no encuentra "Producción"
        ws = wb.active

    stock = {}
    for r in range(2, ws.max_row + 1):
        nombre = _s(ws.cell(r, 1).value).strip()
        cantidad = ws.cell(r, 2).value
        if not nombre:
            continue
        # cantidad puede ser 0 → igual la guardamos para que aparezca en el inventario
        try:
            cantidad_num = float(cantidad) if cantidad not in (None, '') else 0
        except (TypeError, ValueError):
            cantidad_num = 0
        stock[nombre] = cantidad_num
    return stock


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DE LA HOJA NUEVA
# ─────────────────────────────────────────────────────────────────────────────

def generar_output(inv_bytes, mes_objetivo, anio, pool, stock):
    """Modifica el workbook INVENTARIO agregando la hoja del mes_objetivo.
    Retorna los bytes del archivo modificado + estadísticas.
    """
    # Cargamos con formulas (data_only=False) para preservar fórmulas y estilos.
    wb_form = openpyxl.load_workbook(io.BytesIO(inv_bytes), data_only=False)
    # Y una copia con valores calculados para leer INV FINAL del mes anterior.
    wb_val  = openpyxl.load_workbook(io.BytesIO(inv_bytes), data_only=True)

    # Validaciones
    hojas_upper = [sn.strip().upper() for sn in wb_form.sheetnames]
    if mes_objetivo in hojas_upper:
        raise ValueError(
            f'La hoja "{mes_objetivo}" ya existe en el excel INVENTARIO. '
            'Bórrala manualmente antes de reprocesar el mes.'
        )

    # Buscar mes anterior existente (el último mes previo al objetivo dentro de MESES)
    idx_obj = MESES.index(mes_objetivo)
    mes_anterior = None
    for i in range(idx_obj - 1, -1, -1):
        if MESES[i] in hojas_upper:
            mes_anterior = MESES[i]
            break
    if mes_anterior is None:
        raise ValueError(
            f'No encontré ningún mes anterior a {mes_objetivo} en el excel INVENTARIO. '
            'Necesito la hoja del mes anterior para arrancar el nuevo.'
        )

    # Nombres reales (con case original) para acceder a las hojas
    def nombre_real(mes):
        for sn in wb_form.sheetnames:
            if sn.strip().upper() == mes:
                return sn
        return mes

    nombre_real_anterior = nombre_real(mes_anterior)

    # Duplicar hoja del mes anterior
    ws_prev_val  = wb_val[nombre_real_anterior]
    ws_new       = wb_form.copy_worksheet(wb_form[nombre_real_anterior])
    ws_new.title = mes_objetivo

    # Título en A1
    ws_new.cell(1, 1).value = f'INVENTARIO {mes_objetivo}'

    # Recorrer filas de datos: desde 3 hasta max_row de la hoja duplicada
    # Determinar hasta dónde hay data en el mes anterior (por columna A)
    last_data_row = 2
    for r in range(3, ws_prev_val.max_row + 1):
        if ws_prev_val.cell(r, COL_NOMBRE).value not in (None, ''):
            last_data_row = r

    compras_ok = 0
    cantidad_ok = 0

    # Copias mutables del pool y stock (para ir quitando lo matcheado)
    pool_mut  = dict(pool)
    stock_mut = dict(stock)

    for r in range(3, last_data_row + 1):
        nombre_cell = ws_new.cell(r, COL_NOMBRE)
        nombre_raw  = _s(nombre_cell.value)
        nombre_key  = nombre_raw.strip()

        # INV INICIAL ← INV FINAL del mes anterior (valor calculado)
        inv_final_prev = ws_prev_val.cell(r, COL_INV_FINAL).value
        ws_new.cell(r, COL_INV_INICIAL).value = inv_final_prev if inv_final_prev not in (None, '') else 0

        # COMPRAS: limpia primero, luego busca match en pool
        ws_new.cell(r, COL_COMPRAS).value = None
        if nombre_key and nombre_key in pool_mut:
            ws_new.cell(r, COL_COMPRAS).value = pool_mut.pop(nombre_key)
            compras_ok += 1

        # UNITARIO: se conserva (copy_worksheet ya lo trajo, no lo tocamos)

        # CANTIDAD: limpia primero, luego busca match en stock
        ws_new.cell(r, COL_CANTIDAD).value = None
        if nombre_key and nombre_key in stock_mut:
            ws_new.cell(r, COL_CANTIDAD).value = stock_mut.pop(nombre_key)
            cantidad_ok += 1

        # INV FINAL: fórmula
        ws_new.cell(r, COL_INV_FINAL).value = f'=+D{r}*E{r}'

        # COSTEO: fórmula
        ws_new.cell(r, COL_COSTEO).value = f'=+B{r}+C{r}-F{r}'

    # Limpiar celdas por si copy_worksheet arrastró contenido residual en el bloque de sobrantes
    for r in range(2, ws_new.max_row + 1):
        for c in range(COL_SOB_NOMBRE, COL_SOB_CANTIDAD + 1):
            ws_new.cell(r, c).value = None

    # Escribir bloque de sobrantes en col I-K
    # Combinar: unión de nombres pendientes de pool y stock
    todos_sobrantes = sorted(set(pool_mut.keys()) | set(stock_mut.keys()))

    if todos_sobrantes:
        # Headers en fila 2
        header_font = Font(name=FUENTE, bold=True, size=12, color=COLOR_HEADER_SOB_FG)
        header_fill = PatternFill(start_color=COLOR_HEADER_SOB_BG, end_color=COLOR_HEADER_SOB_BG, fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        headers = [
            (COL_SOB_NOMBRE,   'NOMBRE PENDIENTE'),
            (COL_SOB_COMPRAS,  'COMPRAS (pool)'),
            (COL_SOB_CANTIDAD, 'CANTIDAD (stock)'),
        ]
        for col, txt in headers:
            c = ws_new.cell(2, col, value=txt)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align

        # Filas de datos desde fila 3
        data_font = Font(name=FUENTE, size=12)
        for i, nombre in enumerate(todos_sobrantes, start=3):
            ws_new.cell(i, COL_SOB_NOMBRE,   value=nombre).font = data_font
            v_comp = pool_mut.get(nombre)
            v_cant = stock_mut.get(nombre)
            if v_comp is not None:
                ws_new.cell(i, COL_SOB_COMPRAS,  value=v_comp).font = data_font
            if v_cant is not None:
                ws_new.cell(i, COL_SOB_CANTIDAD, value=v_cant).font = data_font

        # Ancho de columnas del bloque de sobrantes
        ws_new.column_dimensions['I'].width = 40
        ws_new.column_dimensions['J'].width = 18
        ws_new.column_dimensions['K'].width = 18

    buf = io.BytesIO()
    wb_form.save(buf)
    buf.seek(0)

    stats = {
        'mes':                  f'{mes_objetivo} {anio}',
        'compras_ok':           compras_ok,
        'compras_pendientes':   len(pool_mut),
        'cantidad_ok':          cantidad_ok,
        'cantidad_pendientes':  len(stock_mut),
    }
    return buf.read(), stats


# ─────────────────────────────────────────────────────────────────────────────
# HANDLER HTTP (Vercel serverless)
# ─────────────────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
            data   = json.loads(body)

            campos = ['archivo_a', 'archivo_b', 'archivo_c']
            faltantes = [c for c in campos if c not in data]
            if faltantes:
                return self._error(400, f'Faltan campos en el request: {faltantes}')

            cargados = [(c, base64.b64decode(data[c])) for c in campos]

            archivos = detectar_archivos(cargados)

            mes_objetivo, anio, _dia = extraer_mes_desde_stock(archivos['stock'])
            pool  = construir_pool(archivos['polizas'])
            stock = construir_stock(archivos['stock'])

            output_bytes, stats = generar_output(
                archivos['inventario'], mes_objetivo, anio, pool, stock
            )

            filename = f'INVENTARIO {anio}.xlsx'

            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('X-Mes',                 stats['mes'])
            self.send_header('X-Compras-Ok',          str(stats['compras_ok']))
            self.send_header('X-Compras-Pendientes',  str(stats['compras_pendientes']))
            self.send_header('X-Cantidad-Ok',         str(stats['cantidad_ok']))
            self.send_header('X-Cantidad-Pendientes', str(stats['cantidad_pendientes']))
            self._cors_headers()
            self.send_header('Access-Control-Expose-Headers',
                'X-Mes, X-Compras-Ok, X-Compras-Pendientes, X-Cantidad-Ok, X-Cantidad-Pendientes')
            self.end_headers()
            self.wfile.write(output_bytes)

        except json.JSONDecodeError:
            self._error(400, 'El cuerpo de la solicitud no es JSON válido')
        except ValueError as e:
            self._error(400, str(e))
        except Exception as e:
            self._error(500, str(e))

    def _error(self, code, message):
        body = json.dumps({'error': message}).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self._cors_headers()
        self.end_headers()
        self.wfile.write(body)
