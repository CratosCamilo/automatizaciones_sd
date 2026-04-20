"""
Vercel Python serverless function — Organización de movimientos bancarios.
Recibe dos Excel (.xlsx) en base64 y un rango de fechas.
Devuelve un Excel de dos hojas (Caja Social + Bancolombia) organizado por día.
"""

from http.server import BaseHTTPRequestHandler
import base64
import io
import json
from collections import defaultdict
from datetime import datetime, date

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

FUENTE     = 'Trebuchet MS'
SIZE_BODY  = 12
SIZE_TOTAL = 20
FMT_NUM    = '_-* #,##0.00_-;\\-* #,##0.00_-;_-* "-"??_-;_-@_-'

BC_EXCLUIR = {'ABONO INTERESES AHORROS', 'IMPTO GOBIERNO 4X1000'}


# ─────────────────────────────────────────────────────────────────────────────
# DETECCIÓN
# ─────────────────────────────────────────────────────────────────────────────

def _detectar_banco(wb):
    if 'AccountMovementsExtended' in wb.sheetnames:
        return 'cajasocial'
    if 'Hoja 1' in wb.sheetnames:
        return 'bancolombia'
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
        for cell in row:
            if cell and 'Titular' in str(cell):
                return 'cajasocial'
    return 'bancolombia'


# ─────────────────────────────────────────────────────────────────────────────
# PARSERS CAJA SOCIAL
# ─────────────────────────────────────────────────────────────────────────────

def _parse_fecha_cs(val):
    if not val:
        return None
    try:
        return datetime.strptime(str(val).strip(), '%d/%m/%Y').date()
    except ValueError:
        return None


def _parse_monto_cs(val):
    if not val or str(val).strip() in ('--', ''):
        return 0
    try:
        return int(round(float(str(val).strip().replace('.', '').replace(',', '.'))))
    except ValueError:
        return 0


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO
# ─────────────────────────────────────────────────────────────────────────────

def procesar_cajasocial(wb, fecha_ini, fecha_fin):
    ws = wb['AccountMovementsExtended']
    rows = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        fecha_str  = row[1]   # col B
        debito_str = row[5]   # col F
        credito_str = row[6]  # col G

        if not fecha_str:
            continue
        fecha = _parse_fecha_cs(fecha_str)
        if not fecha or fecha < fecha_ini or fecha > fecha_fin:
            continue

        debito  = _parse_monto_cs(debito_str)
        credito = _parse_monto_cs(credito_str)
        if debito == 0 and credito == 0:
            continue

        rows.append({
            'fecha':      fecha,
            'fecha_str':  str(fecha_str).strip(),
            'debito':     debito,
            'credito':    credito,
        })
    return rows


def procesar_bancolombia(wb, fecha_ini, fecha_fin):
    ws = wb['Hoja 1']
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fecha_val, desc, ref, valor = row[0], row[1], row[2], row[3]

        if not fecha_val or not desc:
            continue

        if isinstance(fecha_val, datetime):
            fecha = fecha_val.date()
        elif isinstance(fecha_val, date):
            fecha = fecha_val
        else:
            try:
                fecha = datetime.strptime(str(fecha_val).strip(), '%Y-%m-%d').date()
            except (ValueError, AttributeError):
                continue

        if fecha < fecha_ini or fecha > fecha_fin:
            continue

        desc_clean = str(desc).strip()
        if desc_clean.upper() in BC_EXCLUIR:
            continue

        try:
            v = float(valor) if valor is not None else 0.0
        except (ValueError, TypeError):
            v = 0.0

        valor_final = int(v) if isinstance(v, float) and v.is_integer() else v

        rows.append({
            'fecha': fecha,
            'desc':  desc_clean,
            'ref':   str(ref).strip() if ref else '',
            'valor': valor_final,
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# ESCRITURA DE HOJAS
# ─────────────────────────────────────────────────────────────────────────────

def _font(bold=False, size=SIZE_BODY):
    return Font(name=FUENTE, bold=bold, size=size)


def _gap(i, total):
    """Filas vacías entre días: 4 tras el primero, 3 entre el resto, 0 al final."""
    if i >= total - 1:
        return 0
    return 4 if i == 0 else 3


def escribir_hoja_cs(ws, rows):
    ws.cell(row=1, column=1, value='Fecha').font  = _font(bold=True)
    ws.cell(row=1, column=5, value='Débito').font  = _font(bold=True)
    ws.cell(row=1, column=6, value='Crédito').font = _font(bold=True)

    days = defaultdict(list)
    for r in rows:
        days[r['fecha']].append(r)
    sorted_days = sorted(days.keys())

    cur = 2
    for i, day in enumerate(sorted_days):
        day_rows  = sorted(days[day], key=lambda x: x['credito'])
        day_start = cur

        for r in day_rows:
            ws.cell(row=cur, column=1, value=r['fecha_str']).font = _font()

            cell_d = ws.cell(row=cur, column=5, value=r['debito'] if r['debito'] else None)
            cell_d.font = _font()
            if r['debito']:
                cell_d.number_format = FMT_NUM

            cell_c = ws.cell(row=cur, column=6, value=r['credito'] if r['credito'] else None)
            cell_c.font = _font()
            if r['credito']:
                cell_c.number_format = FMT_NUM

            cur += 1

        day_end = cur - 1
        cell_t = ws.cell(row=cur, column=6, value=f'=SUM(F{day_start}:F{day_end})')
        cell_t.font = _font(bold=True, size=SIZE_TOTAL)
        cell_t.number_format = FMT_NUM
        cur += 1

        cur += _gap(i, len(sorted_days))

    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16


def escribir_hoja_bc(ws, rows):
    fill_header = PatternFill(start_color='F2F2F4', end_color='F2F2F4', fill_type='solid')
    for col, name in enumerate(['Fecha', 'Descripción', 'Referencia', 'Valor'], start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font      = _font(bold=True)
        c.fill      = fill_header
        c.alignment = Alignment(horizontal='center')

    days = defaultdict(list)
    for r in rows:
        days[r['fecha']].append(r)
    sorted_days = sorted(days.keys())

    cur = 2
    for i, day in enumerate(sorted_days):
        day_rows = days[day]
        neg = sorted([r for r in day_rows if r['valor'] < 0], key=lambda x: x['valor'])
        pos = sorted([r for r in day_rows if r['valor'] >= 0], key=lambda x: x['valor'])

        first_pos_row = cur + len(neg)

        for r in neg + pos:
            cell_f = ws.cell(row=cur, column=1, value=r['fecha'])
            cell_f.font          = _font()
            cell_f.number_format = 'mm-dd-yy'

            ws.cell(row=cur, column=2, value=r['desc']).font = _font()
            ws.cell(row=cur, column=3, value=r['ref']).font  = _font()

            cell_v = ws.cell(row=cur, column=4, value=r['valor'])
            cell_v.font          = _font()
            cell_v.number_format = FMT_NUM
            cur += 1

        last_row = cur - 1
        if pos:
            total_val = f'=SUM(D{first_pos_row}:D{last_row})'
        else:
            total_val = 0

        cell_t = ws.cell(row=cur, column=4, value=total_val)
        cell_t.font          = _font(bold=True, size=SIZE_TOTAL)
        cell_t.number_format = FMT_NUM
        cur += 1

        cur += _gap(i, len(sorted_days))

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 16


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DEL EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def generar_excel(rows_cs, rows_bc):
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = 'Caja Social'
    escribir_hoja_cs(ws1, rows_cs)

    ws2 = wb.create_sheet('Bancolombia')
    escribir_hoja_bc(ws2, rows_bc)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# HANDLER HTTP (Vercel serverless)
# ─────────────────────────────────────────────────────────────────────────────

MESES = {1:'ENERO',2:'FEBRERO',3:'MARZO',4:'ABRIL',5:'MAYO',6:'JUNIO',
          7:'JULIO',8:'AGOSTO',9:'SEPTIEMBRE',10:'OCTUBRE',11:'NOVIEMBRE',12:'DICIEMBRE'}


class handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin',  '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
            data   = json.loads(body)

            for field in ('archivo1', 'archivo2', 'fecha_inicio', 'fecha_fin'):
                if field not in data:
                    return self._error(400, f'Falta el campo "{field}"')

            fecha_ini = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(data['fecha_fin'],    '%Y-%m-%d').date()

            wb1 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data['archivo1'])))
            wb2 = openpyxl.load_workbook(io.BytesIO(base64.b64decode(data['archivo2'])))

            banco1 = _detectar_banco(wb1)
            banco2 = _detectar_banco(wb2)

            if   banco1 == 'cajasocial'   and banco2 == 'bancolombia':
                wb_cs, wb_bc = wb1, wb2
            elif banco1 == 'bancolombia'  and banco2 == 'cajasocial':
                wb_cs, wb_bc = wb2, wb1
            else:
                return self._error(400,
                    f'No se identificaron los dos bancos. '
                    f'Detectado: archivo1={banco1}, archivo2={banco2}')

            rows_cs = procesar_cajasocial(wb_cs, fecha_ini, fecha_fin)
            rows_bc = procesar_bancolombia(wb_bc, fecha_ini, fecha_fin)
            output  = generar_excel(rows_cs, rows_bc)

            mes = MESES.get(fecha_ini.month, str(fecha_ini.month))
            filename = f'SEMANA {fecha_ini.day} AL {fecha_fin.day} {mes} {fecha_ini.year}.xlsx'

            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('X-CS-Rows', str(len(rows_cs)))
            self.send_header('X-BC-Rows', str(len(rows_bc)))
            self._cors_headers()
            self.send_header('Access-Control-Expose-Headers', 'X-CS-Rows, X-BC-Rows')
            self.end_headers()
            self.wfile.write(output)

        except json.JSONDecodeError:
            self._error(400, 'El cuerpo de la solicitud no es JSON válido')
        except Exception as e:
            self._error(500, str(e))

    def _error(self, code, message):
        body = json.dumps({'error': message}).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self._cors_headers()
        self.end_headers()
        self.wfile.write(body)
