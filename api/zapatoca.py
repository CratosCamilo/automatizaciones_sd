"""
Vercel Python serverless function — Conciliación DIAN vs Zapatoca (inventario).
Recibe los archivos en base64 vía JSON POST y devuelve el Excel procesado.
"""

from http.server import BaseHTTPRequestHandler
import base64
import io
import json
import os
import re
import urllib.request
import zipfile

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────────────

DIAN_KEEP = [
    'Tipo de documento',
    'CUFE/CUDE',
    'Folio',
    'Prefijo',
    'Fecha Emisión',
    'Fecha Recepción',
    'NIT Emisor',
    'Nombre Emisor',
    'IVA',
    'Total',
]

# Proveedores que se omiten en el cruce. Si aparecen en la DIAN no se
# muestran como "falta en sistema"; si aparecen en Zapatoca no se muestran
# como "extra en sistema". Son consumos del negocio (servicios, mercados
# personales, etc.) que no pasan por el inventario.
#
# Esta lista es el SEED / FALLBACK. La lista efectiva se lee del Blob
# público en cada request (ver `cargar_proveedores_excluidos` abajo).
# Slendy edita la lista desde la UI del módulo (icono de engranaje) y se
# persiste en Vercel Blob (`config/proveedores.json`).
DEFAULT_PROVEEDORES_EXCLUIDOS = [
    ('900039901',  'ENERTOTAL S.A. E.S.P.'),
    ('39028745',   'ACEVEDO DE PINILLA LEONOR'),
    ('830055643',  'CINEMARK COLOMBIA S.A.S.'),
    ('901300741',  'COESCO COLOMBIA SAS'),
    ('830122566',  'COLOMBIA TELECOMUNICACIONES S.A. E.S.P. BIC'),
    ('800153993',  'COMUNICACION CELULAR S A  COMCEL S A'),
    ('901419009',  'CR VET SAS'),
    ('890903858',  'INDUSTRIA NACIONAL DE GASEOSAS S.A.S.'),
    ('900280342',  'DISTRIBUCIONES HICAR SAS'),
    ('901893895',  'LA CERVECERIA EXPRESS BODEGA'),
    ('900648058',  'LONG HANG S.A.S.'),
    ('901137699',  'MINISO COLOMBIA S.A.S'),
    ('830037946',  'PANAMERICANA LIBRERÍA Y PAPELERÍA S.A.'),
    ('890903939',  'POSTOBON S.A.'),
    ('830112317',  'PROCAFECOL S.A.'),
    ('900843898',  'RAPPI S.A.S'),
    ('800242106',  'SODIMAC COLOMBIA S.A.'),
    ('900777063',  'SPORTY CITY S.A.S.'),
    ('890100577',  'AEROVIAS DEL CONTINENTE AMERICANO, S.A. AVIANCA - COLOMBIA'),
    ('800216499',  'AGOFER S.A.S.'),
    ('860027404',  'ALLIANZ SEGUROS DE VIDA S A'),
    ('901275662',  'ALMACEN Y TALLER TODO NISSAN A&T S.A.S.'),
    ('900366586',  'ASOCIACION UNION NACIONAL DE COMERCIANTES DE SANTA MARTA'),
    ('901724890',  'C&C INVERSIONES SURTIPROYECTOS S.A.S.'),
    ('901707429',  'CACHARRERIA Y PAPELERÍA CLAUDIA LA 14'),
    ('57270083',   'DIENITH MARIA OROZCO FONSECA'),
    ('890101691',  'GASES DEL CARIBE S A EMPRESA DE SERVICIOS PUBLICOS GASCARIBE S A E S P'),
    ('901132430',  'HIDRODINAMICA SM S.A.S'),
    ('800251569',  'INTER RAPIDISIMO S.A'),
    ('901510218',  'inversiones GA distribuciones sas'),
    ('891700842',  'JARDINES DE PAZ DE SANTA MARTA LTDA'),
    ('1081794799', 'JULYS ROCIO GUTIERREZ CABALLERO'),
    ('91002437',   'PEDRO AMESQUITA BENITEZ'),
    ('1193225562', 'MARIA  JOSE  OLARTE LACERA '),
    ('900167723',  'MULTIDESARROLLOS URBANOS S.A.S.'),
    ('900834997',  'PROQUIMAG SAS'),
    ('900031973',  'PUERTO DIESEL LTDA'),
    ('63506621',   'RUBIELA GOMEZ MORENO'),
    ('830512656',  'PRODUCTOS INDUSTRIALES Y ASESORIAS S.A.S. PROINAS S.A.S'),
    ('900724151',  'SOLAB SAS'),
    ('901433765',  'SOLO RODAMIENTOS Y RETENES SAS'),
    ('800027374',  'TECNOLOGIA ALIMENTARIA S.A.S. BIC'),
    ('57428408',   'TIANA ISABEL MARCUCCI BECERRA'),
    ('900092385',  'UNE EPM TELECOMUNICACIONES S.A'),
    ('891780160',  'CAMARA DE COMERCIO DE SANTA MARTA PARA EL MAGDALENA'),
    ('901916738',  'INDUSTRIAS WESCOLD OUTLET SAS'),
    ('900486370',  'YOYO S.A.S.'),
    ('901798108',  'COMERCIALIZADORA INTEGRAL DE TECNOLOGIA S.A.S.'),
    ('890200928',  'COOPERATIVA SANTANDEREANA DE TRANSPORTADORES LIMITADA'),
    ('819001879',  'BASCULAS Y BALANZAS DE LA COSTA LTDA'),
    ('92517327',   'ALIRIO DANILO OLIVERA CORREA'),
    ('890900943',  'COLOMBIANA DE COMERCIO S.A.'),
    ('901229027',  'ECOLPLAGAS S.A.S'),
    ('860008424',  'HANSEATICA SAS'),
    ('860037013',  'COMPAÑIA MUNDIAL DE SEGUROS S.A.'),
    ('900375008',  'PROVINAS SAS'),
    ('901214287',  'SOINGTEL'),
    ('900543551',  'CORREAS GUAYAS Y ENSAMBLES S.A.S.'),
    ('800144352',  'HIDROANDINA Y COMPAÑIA LIMITADA'),
    ('860009578',  'SEGUROS DEL ESTADO S.A.'),
    ('900430690',  'FIRENZECORP S.A.S'),
    ('1128191884', 'HERRAMIENTAS EXPRESS'),
    ('800185781',  'AEROREPUBLICA S.A.'),
    ('800242394',  'ALMACEN REFRIELECTRIC S.A.S'),
    ('890929497',  'MAPER S.A.S'),
    ('860003981',  'COLMAQUINAS S.A.'),
]


# ─────────────────────────────────────────────────────────────────────────────
# LECTURA DINÁMICA DE LA LISTA DESDE VERCEL BLOB
# ─────────────────────────────────────────────────────────────────────────────

BLOB_PATHNAME_PROVEEDORES = 'config/proveedores.json'


def _blob_base_url():
    """Deriva la URL pública del Blob store desde el RW token.
    Token formato: 'vercel_blob_rw_<STORE_ID>_<random>'."""
    token = os.environ.get('BLOB_READ_WRITE_TOKEN', '')
    parts = token.split('_')
    if len(parts) < 5 or parts[0] != 'vercel' or parts[1] != 'blob':
        return None
    return f'https://{parts[3].lower()}.public.blob.vercel-storage.com'


def cargar_proveedores_excluidos():
    """Devuelve la lista efectiva de proveedores excluidos.
    Intenta leer del Blob; si falla o está vacío, cae a los defaults.
    Formato interno: lista de tuplas (nit, nombre)."""
    base = _blob_base_url()
    if base:
        try:
            import time
            url = f'{base}/{BLOB_PATHNAME_PROVEEDORES}?t={time.time_ns()}'
            req = urllib.request.Request(url, headers={'Cache-Control': 'no-cache'})
            with urllib.request.urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode('utf-8'))
            if isinstance(data, list) and data:
                out = []
                for item in data:
                    nit = str(item.get('nit', '')).strip()
                    nom = str(item.get('nombre', '')).strip()
                    if nit and nom:
                        out.append((nit, nom))
                if out:
                    return out
        except Exception:
            pass
    return DEFAULT_PROVEEDORES_EXCLUIDOS

COLS_MONETARIAS_DIAN = {'IVA', 'Total'}
FMT_MILES   = '#,##0.##'
FUENTE      = 'Trebuchet MS'
OUTPUT_NAME = 'CONCILIACION_INVENTARIO_VS_DIAN.xlsx'

# Colores (matchean el ejemplo entregado por Slendy)
COLOR_HEADER_DIAN_BG = '348441'  # verde oscuro
COLOR_HEADER_ZAP_BG  = '007D7E'  # teal
COLOR_HEADER_FG      = 'FFFFFF'
COLOR_FALTA_BG       = 'F2DCDB'  # rojo claro — en DIAN, no en Zapatoca
COLOR_EXTRA_BG       = 'FDE9D9'  # naranja claro — en Zapatoca, no en DIAN
COLOR_OK_BG          = 'EBF1DE'  # verde claro — matcheadas
COLOR_TEXTO          = '000000'
COLOR_BORDE          = 'BFBFBF'


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _borde():
    s = Side(style='thin', color=COLOR_BORDE)
    return Border(left=s, right=s, top=s, bottom=s)


def _clean(v):
    if v is None:
        return None
    if isinstance(v, float) and v != v:  # NaN
        return None
    return v


def _normalize_key(s):
    """Normaliza una clave de cruce: quita todo lo no alfanumérico y pasa a mayúsculas."""
    if s is None:
        return ''
    return re.sub(r'[^A-Za-z0-9]', '', str(s)).upper()


def _folio_str(folio):
    """Convierte un folio numérico (98022.0) a string limpio ('98022')."""
    if folio is None:
        return ''
    if isinstance(folio, float) and folio.is_integer():
        return str(int(folio))
    return str(folio).strip()


def _folio_solo(zap_folio):
    """De un folio Zapatoca como 'FEIP-98022' o 'BBQR 168170' extrae el último
    grupo de dígitos ('98022', '168170'). Útil cuando la DIAN trae prefijo
    vacío y el cruce por prefijo+folio falla.
    """
    if zap_folio is None:
        return ''
    m = re.search(r'(\d+)\s*$', str(zap_folio))
    return m.group(1) if m else ''


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO DIAN
# ─────────────────────────────────────────────────────────────────────────────

def procesar_dian(zip_bytes, nits_excluidos):
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith('.xlsx')]
        if not xlsx_names:
            raise ValueError('No se encontró ningún .xlsx dentro del ZIP de la DIAN')
        xlsx_data = zf.read(xlsx_names[0])

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    faltantes = [c for c in DIAN_KEEP if c not in headers]
    if faltantes:
        raise ValueError(
            f'DIAN: columnas faltantes {faltantes}. Disponibles: {headers}'
        )

    idx = {col: headers.index(col) for col in DIAN_KEEP}

    filas = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None:
            continue
        fila = {col: raw[idx[col]] for col in DIAN_KEEP}

        total = fila['Total']
        try:
            total_num = float(total) if total not in (None, '') else 0
        except (TypeError, ValueError):
            total_num = 0
        if total_num == 0:
            continue
        fila['Total'] = total_num

        try:
            fila['IVA'] = float(fila['IVA']) if fila['IVA'] not in (None, '') else 0
        except (TypeError, ValueError):
            fila['IVA'] = 0

        nit = str(fila['NIT Emisor']).strip() if fila['NIT Emisor'] is not None else ''
        fila['NIT Emisor'] = nit

        if nit in nits_excluidos:
            continue

        fila['_key'] = _normalize_key(fila['Prefijo']) + _normalize_key(_folio_str(fila['Folio']))
        filas.append(fila)

    return filas


# ─────────────────────────────────────────────────────────────────────────────
# PROCESAMIENTO ZAPATOCA
# ─────────────────────────────────────────────────────────────────────────────

ZAP_REQUIRED = ['Fecha', 'Folio', 'Proveedor', 'Total']


def procesar_zapatoca(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    faltantes = [c for c in ZAP_REQUIRED if c not in headers]
    if faltantes:
        raise ValueError(
            f'Zapatoca: columnas faltantes {faltantes}. Disponibles: {headers}'
        )

    idx = {col: headers.index(col) for col in ZAP_REQUIRED}

    filas = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw is None or all(v is None for v in raw):
            continue
        total = raw[idx['Total']]
        try:
            total_num = float(total) if total not in (None, '') else 0
        except (TypeError, ValueError):
            total_num = 0
        if total_num == 0:
            continue

        folio_raw = raw[idx['Folio']]
        fecha_raw = raw[idx['Fecha']]
        proveedor = raw[idx['Proveedor']]

        filas.append({
            'Fecha':     fecha_raw,
            'Folio':     folio_raw,
            'Proveedor': proveedor,
            'Total':     total_num,
            '_key':      _normalize_key(folio_raw),
            '_folio':    _folio_solo(folio_raw),
        })

    return filas


# ─────────────────────────────────────────────────────────────────────────────
# CRUCE
# ─────────────────────────────────────────────────────────────────────────────

def cruzar(dian_filas, zap_filas):
    """Devuelve (matched, dian_only, zap_only).

    Match en dos pasadas:
      1) clave normalizada PREFIJO+FOLIO contra el folio Zapatoca normalizado.
      2) DIAN con prefijo vacío (folio puro) contra el grupo de dígitos final
         del folio Zapatoca. Cubre proveedores donde la DIAN no reporta
         prefijo pero Zapatoca sí (ej. INSUMO PAN → FEIP-XXXX).
    """
    zap_pool = {}
    for f in zap_filas:
        zap_pool.setdefault(f['_key'], []).append(f)

    matched     = []
    dian_resto  = []

    for d in dian_filas:
        k = d['_key']
        if k and k in zap_pool and zap_pool[k]:
            matched.append((d, zap_pool[k].pop(0)))
        else:
            dian_resto.append(d)

    # Pasada 2: DIAN sin prefijo vs folio-solo de Zapatoca
    zap_pool_folio = {}
    for lst in zap_pool.values():
        for z in lst:
            zap_pool_folio.setdefault(z['_folio'], []).append(z)

    dian_only = []
    for d in dian_resto:
        prefijo = _normalize_key(d['Prefijo'])
        folio   = _normalize_key(_folio_str(d['Folio']))
        if not prefijo and folio and folio in zap_pool_folio and zap_pool_folio[folio]:
            z = zap_pool_folio[folio].pop(0)
            # también lo saco del pool principal
            zap_pool[z['_key']].remove(z)
            matched.append((d, z))
        else:
            dian_only.append(d)

    zap_only = [z for lst in zap_pool.values() for z in lst]

    matched.sort(key=lambda dz: dz[0]['Total'], reverse=True)
    dian_only.sort(key=lambda d: d['Total'], reverse=True)
    zap_only.sort(key=lambda z: z['Total'], reverse=True)

    return matched, dian_only, zap_only


# ─────────────────────────────────────────────────────────────────────────────
# GENERACIÓN DEL EXCEL
# ─────────────────────────────────────────────────────────────────────────────

DIAN_HEADERS_OUT = DIAN_KEEP[:]      # cols 1..10
DIFF_HEADER      = 'DIFERENCIA'      # col 11
ZAP_HEADERS_OUT  = ['Total', 'Fecha', 'Folio', 'Proveedor']  # cols 12..15

# Anchos según ejemplo
COLUMN_WIDTHS = {
    'A': 29.3, 'B': 13.0, 'C': 14.1, 'D': 13.3,
    'E': 21.6, 'F': 24.4, 'G': 17.6, 'H': 56.3,
    'I': 13.6, 'J': 15.6, 'K': 15.6, 'L': 15.6,
    'M': 14.1, 'N': 15.6, 'O': 53.1,
}


def _escribir_fila_dian(ws, r, d):
    for c, col in enumerate(DIAN_HEADERS_OUT, start=1):
        ws.cell(row=r, column=c, value=_clean(d[col]))


def _escribir_fila_zap(ws, r, z):
    ws.cell(row=r, column=12, value=_clean(z['Total']))
    ws.cell(row=r, column=13, value=_clean(z['Fecha']))
    ws.cell(row=r, column=14, value=_clean(z['Folio']))
    ws.cell(row=r, column=15, value=_clean(z['Proveedor']))


def generar_output(matched, dian_only, zap_only, proveedores_excluidos):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hoja1'

    # Encabezados
    for c, col in enumerate(DIAN_HEADERS_OUT, start=1):
        ws.cell(row=1, column=c, value=col)
    ws.cell(row=1, column=11, value=DIFF_HEADER)
    for c, col in enumerate(ZAP_HEADERS_OUT, start=12):
        ws.cell(row=1, column=c, value=col)

    r = 2

    # 1) DIAN-only (rojo): en DIAN, falta en sistema
    fila_red_ini = r
    for d in dian_only:
        _escribir_fila_dian(ws, r, d)
        r += 1
    fila_red_fin = r - 1

    # 2) Zap-only (naranja): en sistema, no en DIAN
    fila_extra_ini = r
    for z in zap_only:
        _escribir_fila_zap(ws, r, z)
        r += 1
    fila_extra_fin = r - 1

    # 3) Matched (verde): en ambos
    fila_ok_ini = r
    for d, z in matched:
        _escribir_fila_dian(ws, r, d)
        ws.cell(row=r, column=11, value=f'=+J{r}-L{r}')
        _escribir_fila_zap(ws, r, z)
        r += 1
    fila_ok_fin = r - 1

    fila_final = r - 1

    # ── Estilos ───────────────────────────────────────────────────────────
    borde = _borde()

    # Encabezados DIAN (A..J): fondo verde oscuro
    fill_h_dian = _fill(COLOR_HEADER_DIAN_BG)
    font_h_dian = Font(name=FUENTE, bold=True, color=COLOR_HEADER_FG, size=12)
    for c in range(1, 11):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill_h_dian
        cell.font = font_h_dian
        cell.border = borde
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Encabezado DIFERENCIA (K): sin fondo, negrilla
    cellK = ws.cell(row=1, column=11)
    cellK.font = Font(name=FUENTE, bold=True, color=COLOR_TEXTO, size=12)
    cellK.border = borde
    cellK.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Encabezados Zapatoca (L..O): fondo teal
    fill_h_zap = _fill(COLOR_HEADER_ZAP_BG)
    font_h_zap = Font(name=FUENTE, bold=True, color=COLOR_HEADER_FG, size=12)
    for c in range(12, 16):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill_h_zap
        cell.font = font_h_zap
        cell.border = borde
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Filas de datos: color de fondo por grupo
    font_data = Font(name=FUENTE, color=COLOR_TEXTO, size=11)
    cols_monetarias = {9, 10, 11, 12}  # IVA, Total DIAN, DIFERENCIA, Total Zap

    def pintar(rini, rfin, hex_color):
        if rini > rfin:
            return
        fill = _fill(hex_color)
        for rr in range(rini, rfin + 1):
            for cc in range(1, 16):
                cell = ws.cell(row=rr, column=cc)
                cell.fill = fill
                cell.font = font_data
                cell.border = borde
                if cc in cols_monetarias:
                    cell.number_format = FMT_MILES

    pintar(fila_red_ini,   fila_red_fin,   COLOR_FALTA_BG)
    pintar(fila_extra_ini, fila_extra_fin, COLOR_EXTRA_BG)
    pintar(fila_ok_ini,    fila_ok_fin,    COLOR_OK_BG)

    # Anchos de columna fijos
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 30

    # ── Hoja 2: proveedores excluidos ──────────────────────────────────────
    ws2 = wb.create_sheet('BORRAR ESTOS PROVEEDORES')
    fill_excl = _fill(COLOR_FALTA_BG)
    font_excl = Font(name=FUENTE, color=COLOR_TEXTO, size=11)
    for i, (nit, nombre) in enumerate(proveedores_excluidos, start=1):
        c1 = ws2.cell(row=i, column=1, value=nit)
        c2 = ws2.cell(row=i, column=2, value=nombre)
        for c in (c1, c2):
            c.fill = fill_excl
            c.font = font_excl
            c.border = borde
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 52

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# HANDLER HTTP (Vercel serverless)
# ─────────────────────────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass

    def _cors_headers(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors_headers()
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
            data   = json.loads(body)

            if 'dian' not in data or 'zapatoca' not in data:
                return self._error(400, 'Se requieren los campos "dian" y "zapatoca" en base64')

            zip_bytes  = base64.b64decode(data['dian'])
            xlsx_bytes = base64.b64decode(data['zapatoca'])

            # Lista dinámica de excluidos (leída del Blob en cada request)
            proveedores_excluidos = cargar_proveedores_excluidos()
            nits_excluidos = {nit for nit, _ in proveedores_excluidos}

            dian_filas = procesar_dian(zip_bytes, nits_excluidos)
            zap_filas  = procesar_zapatoca(xlsx_bytes)

            matched, dian_only, zap_only = cruzar(dian_filas, zap_filas)

            output_bytes = generar_output(matched, dian_only, zap_only, proveedores_excluidos)

            self.send_response(200)
            self.send_header('Content-Type',
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition',
                f'attachment; filename="{OUTPUT_NAME}"')
            self.send_header('X-Falta', str(len(dian_only)))
            self.send_header('X-Extra', str(len(zap_only)))
            self.send_header('X-Ok',    str(len(matched)))
            self._cors_headers()
            self.send_header('Access-Control-Expose-Headers', 'X-Falta, X-Extra, X-Ok')
            self.end_headers()
            self.wfile.write(output_bytes)

        except json.JSONDecodeError:
            self._error(400, 'El cuerpo de la solicitud no es JSON válido')
        except Exception as e:
            self._error(500, str(e))

    def _error(self, code, message):
        body = json.dumps({'error': message}).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self._cors_headers()
        self.end_headers()
        self.wfile.write(body)
